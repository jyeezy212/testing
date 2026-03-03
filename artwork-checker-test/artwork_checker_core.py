#!/usr/bin/env python3
"""
================================================================================
ARTWORK CHECKER v2.1.0 - XLSX COPY DOCUMENT VERSION
================================================================================

A comprehensive tool for validating packaging artwork against copy documents.
Designed for use with ChatGPT Custom GPT for visual verification workflow.

INPUT:  Copy document (.xlsx) + Artwork file (.pdf or .ai)
OUTPUT: Markdown verification report

FEATURES:
---------
✅ Excel copy document parsing (openpyxl)
✅ PDF/AI text extraction via PyMuPDF
✅ Character-by-character comparison
✅ Smart zoom triggers for visual verification
✅ Comprehensive exclusion patterns for non-product text
✅ INCI capitalization validation
✅ Barcode scanning and validation
✅ Full-page annotated snapshots
✅ Markdown report generation
✅ Anti-hallucination safeguards
✅ Curved text detection
✅ Underline detection
✅ Cross-reference between Section 3A and 3D findings

USAGE:
------
    python artwork_checker_v2_1_xlsx.py --copy <copy.xlsx> --artwork <artwork.pdf>

AUTHOR: Built for Amika brand artwork verification
VERSION: 2.1.0
DATE: 2025
================================================================================
"""

# =============================================================================
# IMPORTS
# =============================================================================

import logging
import re
import argparse
import sys
import traceback
import json
import os
import csv
import hashlib
import math
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Set, Any, Union
from dataclasses import dataclass, field
from enum import Enum, auto
from datetime import datetime
from difflib import SequenceMatcher
from io import BytesIO

# Third-party imports - with availability checks
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

try:
    from PIL import Image, ImageDraw, ImageFont
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    from pyzbar.pyzbar import decode as decode_barcode
    PYZBAR_AVAILABLE = True
except ImportError:
    PYZBAR_AVAILABLE = False

# =============================================================================
# LOGGING SETUP
# =============================================================================

def setup_logging(level: int = logging.INFO) -> logging.Logger:
    logging.basicConfig(
        level=level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    return logging.getLogger(__name__)

logger = setup_logging()


# =============================================================================
# CUSTOM EXCEPTIONS
# =============================================================================

class VisionPassRequired(Exception):
    """
    Raised by run_check() when Pass 1 completes and visual verification is
    needed before the final report can be generated.

    Attributes
    ----------
    tasks_path   : str  – path to vision_tasks.json
    tier1_count  : int  – number of items requiring visual review
    """

    def __init__(self, tasks_path: str, tier1_count: int, message: str):
        super().__init__(message)
        self.tasks_path = tasks_path
        self.tier1_count = tier1_count


class VisionOverrideRejected(Exception):
    """Raised when vision_overrides.json fails validation (missing audit stamp,
    wrong task hash, suspicious auto-fill, missing evidence, etc.)."""


# =============================================================================
# ENUMERATIONS
# =============================================================================

class ExtractionMethod(Enum):
    LIVE_TEXT = "live_text"
    AI_PDF_COMPATIBLE = "ai_pdf"
    VISION_SYSTEM = "vision"
    FAILED = "failed"


class MatchType(Enum):
    EXACT_MATCH = "exact"
    NEAR_MATCH = "near"
    MISMATCH = "mismatch"
    MISSING_IN_ARTWORK = "missing"
    EXTRA_IN_ARTWORK = "extra"
    REQUIRES_VERIFICATION = "verify"


class StatusCode(Enum):
    OK = "OK"
    ATTN = "ATTN"
    FAIL = "FAIL"
    TBD = "TBD"
    FYI = "FYI"


class RiskLevel(Enum):
    LOW = "Low"
    MEDIUM = "Medium"
    HIGH = "High"


# =============================================================================
# CONFIGURATION
# =============================================================================

class Config:
    """Central configuration for the artwork checker."""

    # VERSION INFO
    VERSION = "2.1.0"
    BUILD_DATE = "2025-06"

    # MATCHING THRESHOLDS
    EXACT_MATCH_THRESHOLD = 100.0
    NEAR_MATCH_THRESHOLD = 95.0
    MISMATCH_THRESHOLD = 95.0

    # ZOOM TRIGGERS
    ZOOM_FONT_SIZE_THRESHOLD = 6.5
    ZOOM_ON_NUMBERS = True
    ZOOM_ON_PERCENTAGE = True
    ZOOM_ON_DECIMALS = True
    ZOOM_ON_UNITS = True
    ZOOM_ON_NEGATION = True
    ZOOM_CONFIDENCE_THRESHOLD = 100
    ZOOM_FUZZY_THRESHOLD = 100
    VISION_ALWAYS_REQUIRED = True

    # VISION OVERRIDE VALIDATION (Pass 2 gate)
    VISION_OVERRIDE_STRICT = True
    VISION_OVERRIDE_REQUIRE_AUDIT_STAMP = True
    VISION_OVERRIDE_REQUIRE_EVIDENCE = True

    # ROTATION RECONSTRUCTION
    ROTATION_PROXIMITY_THRESHOLD = 8.0

    UNIT_PATTERNS = [
        r'\b\d+\s*(mg|g|kg|ml|mL|ML|l|L|oz|OZ|fl\.?\s*oz|FL\.?\s*OZ|pt|qt|gal)\b',
        r'\b\d+\s*(mm|cm|m|in|inch|inches|ft|feet)\b',
        r'\bUS\s+FL\.?\s*OZ\.?\b',
    ]

    NEGATION_WORDS = [
        'no', 'not', 'free', 'only', 'without', 'never', 'none',
        'zero', 'moins', 'sans', 'non', 'aucun'
    ]

    # -------------------------------------------------------------------------
    # EXCEL TEMPLATE STRUCTURE
    # -------------------------------------------------------------------------
    # Row-to-panel mapping (1-indexed to match Excel rows)
    FRONT_PANEL_ROWS = range(2, 6)   # Rows 2-5
    BACK_PANEL_ROWS = range(6, 31)   # Rows 6-30

    # Row 4 = Fill Weight, exempt from lowercase checking
    FILL_WEIGHT_ROW = 4

    # Rows that require lowercase checking
    LOWERCASE_CHECK_ROWS = [2, 3] + list(range(5, 23))  # Rows 2,3,5-22

    # Pack Claim rows for claim risk assessment (rows 11-15)
    PACK_CLAIM_ROWS = range(11, 16)

    # Language column mapping (0-indexed for openpyxl column access)
    LANGUAGE_COLUMNS = {
        2: 'English',      # Column C
        3: 'French',       # Column D
        4: 'Spanish',      # Column E
        5: 'German',       # Column F
        6: 'Dutch',        # Column G
        7: 'Italian',      # Column H
        8: 'Scan/Danish',  # Column I
        9: 'Finnish',      # Column J
        10: 'Portuguese',  # Column K
        11: 'Polish',      # Column L
        12: 'Russian',     # Column M
    }

    # Language prefix abbreviations (must be uppercase)
    LANGUAGE_PREFIXES = {
        'French': 'FR',
        'Spanish': 'ES',
        'German': 'DE',
        'Dutch': 'NL',
        'Italian': 'IT',
        'Scan/Danish': 'DA',
        'Finnish': 'FI',
        'Portuguese': 'PT',
        'Polish': 'PL',
        'Russian': 'RU',
    }

    # Row 16 special label
    HERO_INGREDIENTS_HEADER_ROW = 16

    # Rows that are always checked for instructional content (28, 29, 30)
    INSTRUCTIONAL_CHECK_ROWS = [28, 29, 30]

    # Row 30 = UPC row
    UPC_ROW = 30

    # -------------------------------------------------------------------------
    # FIELDS EXEMPT FROM LOWERCASE CHECKING
    # -------------------------------------------------------------------------
    UPPERCASE_ALLOWED_FIELDS = [
        'Address Block',
        'Biorius Address',
        'Formula Country of Origin',
        'Ingredient List',
        'Fill Weight',
    ]

    ALLOWED_ACRONYMS = [
        'ML', 'mL', 'FL', 'OZ', 'USA', 'EU', 'UK', 'GB', 'BE', 'CA',
        'AHA', 'BHA', 'LHA', 'NP', 'SPF', 'UV', 'UVA', 'UVB',
        'PEG', 'PPG', 'EDTA', 'BHT',
        'NY', 'NYC', 'LA', 'SF',
        'LLC', 'INC', 'CO', 'LTD',
    ]

    INCI_LOWERCASE_CONNECTORS = [
        'de', 'du', 'des', 'la', 'le', 'les', "d'", "l'",
        'of', 'the', 'and', 'with',
        'et', 'cum',
    ]

    # -------------------------------------------------------------------------
    # EXCLUSION PATTERNS (non-product text from artwork)
    # -------------------------------------------------------------------------
    EXCLUSION_EXACT_MATCHES = {
        'white', 'black', 'cyan', 'magenta', 'yellow',
        'c', 'm', 'y', 'k', 'cmyk', 'rgb',
        'spot gloss', 'matte', 'gloss', 'foil',
        'bleed', 'trim', 'safe zone', 'die cut', 'dieline', 'die line',
        'fold', 'cut', 'glue', 'flap', 'score',
        'crop marks', 'fold here', 'cut here', 'glue flap',
        'proof', 'draft', 'final', 'approved',
        'fpo', 'for position only',
        'ol', 'outlined',
        'front', 'back', 'side', 'top', 'bottom',
        'print side', 'inside', 'outside',
        'dieline',
    }

    EXCLUSION_REGEX_PATTERNS = [
        # Dimension patterns
        r'^\d+\+?\d*/\d+$',
        r'^\d+\s+\d+/\d+$',
        r'^\d+\.\d{2,}$',
        r'^\d+-\d+/\d+"?$',
        # Reference numbers
        r'^IP\d+[A-Z]?$',
        r'^VC[_-]?\d+$',
        r'^ref\s*#?\s*:',
        r'^code\s*#?\s*:',
        # Pantone/PMS colors
        r'^PMS\s*\d+',
        r'^Pantone\s*\d+',
        r'^PMS\s+\w+\.\s*\w+',
        r'^pms\s+\d+',
        # Technical specifications
        r'^SCALE\s*:',
        r'^SIZE\s*:',
        r'^DATE\s*:',
        r'^DRAWN\s*:',
        r'^CHECKED\s*:',
        r'^APPROVED\s*:',
        r'^MATERIAL\s*:',
        r'^WEIGHT\s*:',
        r'^DWG\s*NO',
        r'^OFC\s*:',
        r'^UNIT\s*:',
        r'^NAME\s*:',
        r'^DESCRIPTION\s*:',
        # Tolerances
        r'^[X\.]+±\d+',
        r'^\.\w+±\d+',
        r'^X°±\d+',
        r'GENERAL\s+TOLERANCES',
        r'MILLIMETRES',
        r'INCHES',
        # Company boilerplate
        r'INTEGRATED\s+PACKAGING',
        r'Kroger\s+Packaging',
        r'PROPRIETARY',
        r'CONFIDENTIAL',
        r'COPYRIGHT',
        r'This\s+design\s+concept\s+is\s+the\s+exclusive\s+property',
        r'all\s+rights\s+are\s+reserved',
        r'end\s+user\s*:',
        r'item\s*:',
        r'3RD\s+ANGLE\s+PROJECTION',
        # Label/packaging metadata
        r'^Label\s+Dieline',
        r'^Customer\s*:',
        r'^Diameter',
        r'^CLEARANCE\s+AREA',
        r'^Deco\s+Area',
        # Version/date stamps and artwork slug metadata
        r'^\d{1,2}/\d{1,2}/\d{2,4}$',
        r'^\d{4}\.\d{2}\.\d{2}$',
        r'^date\s+revised',
        r'^product\s+name\s+vendor',
        r'^notes$',
        r'^phase$',
        r'final\s+AW$',
        r'^\d+$',  # Bare numbers (dimensions, page numbers)
        # Diameter/size specs
        r'^Ø\d+',
        r'^\d+\s*x\s*\d+',
        # View labels
        r'^view\s*:',
        r'^date\s*:',
        r'^size\s*:',
    ]

    EXCLUSION_COMPANY_NAMES = [
        'INTEGRATED PACKAGING INDUSTRIES',
        'Kroger Packaging Inc',
        'Heat Makes Sense',
        'NVI',
    ]

    # -------------------------------------------------------------------------
    # INSTRUCTIONAL NOTE DETECTION PATTERNS
    # -------------------------------------------------------------------------
    INSTRUCTIONAL_PATTERNS = [
        # Yes/No starters
        r'^yes\s*[-–—]',
        r'^no\s*[-–—]',
        # Pending/placeholder indicators
        r'\bTBD\b',
        r'\bTBC\b',
        r'\bTODO\b',
        r'\bXXX+\b',
        r'\bPLACEHOLDER\b',
        r'\b[Pp]ending\b',
        # Production references
        r'\bPO\b',
        r'\bfirst\s+PO\b',
        r'\bproduction\b',
        # Conditional language
        r'if\s+it\s+can\s+fit',
        r'if\s+it\s+can\s+be\s+fit',
        r'need\s+to\s+wait',
        r'waiting\s+for',
        r'once\s+confirmed',
        # Parenthetical instructions
        r'\(pending\)',
        r'\(see\s+attached\)',
        r'\(if\s+',
        r'\(TBD\)',
        r'\(optional\)',
        # Mentions and assignments
        r'@\w+',
        r'\b(?:ask|check\s+with|confirm\s+with)\b',
        # Date references in instruction context
        r'(?:updated|as\s+of|by)\s+\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',
        # Regulatory/confirmation notes
        r'(?:Reg|regulatory)\s+confirmation',
        r'certification\s+(?:pending|needed)',
        # Cross-reference instructions
        r'details\s+on\s+\w+',
        r'^n/a\s*:',
        r'^n/a$',
        r'^N/A\s*:',
        r'^N/A$',
        r'on\s+unit\s+carton',
        r'^see\s+\w+',
        r'NO\s+UPC\s+NEEDED',
        r'JUST\s+SET\s+BOX',
        r'^Picture$',
    ]

    # -------------------------------------------------------------------------
    # CLAIM RISK TERMS
    # -------------------------------------------------------------------------
    HIGH_RISK_CLAIM_TERMS = [
        'clinically proven', 'clinically tested', 'clinically demonstrated',
        'dermatologist tested', 'dermatologist approved', 'dermatologist recommended',
        'doctor recommended', 'physician recommended',
        'scientifically proven', 'scientifically tested',
        'medically proven', 'FDA approved', 'FDA cleared',
        'patented', 'patent pending',
        'cliniquement prouvé', 'testé cliniquement',
        'dermatologiquement testé',
    ]

    MEDIUM_RISK_CLAIM_TERMS = [
        'reduces', 'eliminates', 'removes', 'prevents',
        'anti-aging', 'anti-wrinkle', 'anti-acne',
        'repairs', 'restores', 'regenerates', 'renews',
        'strengthens', 'fortifies', 'rebuilds',
        'treats', 'heals', 'cures',
        'deeply moisturizes', 'intensely hydrates',
        'réduit', 'élimine', 'prévient',
    ]

    # -------------------------------------------------------------------------
    # BARCODE CONFIGURATION
    # -------------------------------------------------------------------------
    BARCODE_RENDER_DPI = 4
    BARCODE_TYPES = {
        'UPC-A': 12,
        'UPC-E': 8,
        'EAN-13': 13,
        'EAN-8': 8,
    }

    # -------------------------------------------------------------------------
    # SNAPSHOT CONFIGURATION
    # -------------------------------------------------------------------------
    SNAPSHOT_ENABLED = True
    SNAPSHOT_DPI = 150
    SNAPSHOT_MAX_PER_PAGE = 20
    ANNOTATION_COLORS = {
        'FAIL': (255, 0, 0),
        'ATTN': (255, 165, 0),
        'OK': (0, 255, 0),
        'TBD': (128, 128, 128),
    }
    ANNOTATION_LINE_WIDTH = 3

    # -------------------------------------------------------------------------
    # OUTPUT CONFIGURATION
    # -------------------------------------------------------------------------
    STATUS_EMOJI = {
        'OK': '✅',
        'ATTN': '⚠️',
        'FAIL': '❌',
        'TBD': '🔍',
        'FYI': 'ℹ️',
    }

    # -------------------------------------------------------------------------
    # BRAND
    # -------------------------------------------------------------------------
    BRAND_NAME = "amika"


# Global config instance
config = Config()


# =============================================================================
# DATA MODELS
# =============================================================================

@dataclass
class TextRun:
    """A single extracted text element from artwork with metadata."""
    text: str
    page_number: int
    bbox: Optional[Tuple[float, float, float, float]] = None
    font_name: Optional[str] = None
    font_size: Optional[float] = None
    extraction_method: ExtractionMethod = ExtractionMethod.LIVE_TEXT
    confidence: float = 1.0
    rotation: int = 0
    is_underlined: bool = False


@dataclass
class CopyField:
    """A field from the copy document to be verified."""
    field_name: str
    panel: str
    language: str
    text: str
    row_number: int
    col_index: int
    is_instructional: bool = False
    instructional_pattern: Optional[str] = None
    is_underlined: bool = False
    display_name: Optional[str] = None  # Override for report (e.g., "Hero Ingredients Header")


@dataclass
class MatchFinding:
    """Result of comparing a copy field against artwork."""
    field_name: str
    panel: str
    language: str
    copy_value: str
    artwork_value: Optional[str]
    match_type: MatchType
    similarity_score: float
    status_code: StatusCode
    notes: List[str] = field(default_factory=list)
    requires_zoom: bool = False
    zoom_reasons: List[str] = field(default_factory=list)
    issue_id: Optional[str] = None
    bbox: Optional[Tuple[float, float, float, float]] = None
    matched_runs: List['TextRun'] = field(default_factory=list)
    has_3a_flag: bool = False  # Cross-reference: copy doc had quality issue


@dataclass
class CopyQualityIssue:
    """An issue found in the copy document itself (Section 3A)."""
    language: str
    field_name: str
    original_text: str
    issue_type: str
    recommendation: str
    status_code: StatusCode


@dataclass
class ClaimRisk:
    """Risk assessment for a regulatory claim (Section 3B)."""
    language: str
    claim_text: str
    risk_level: RiskLevel
    rationale: str
    regions: List[str]
    recommended_action: str
    status_code: StatusCode


@dataclass
class ConversionCheck:
    """Volume/weight conversion verification (Section 3C)."""
    source_field: str
    declared_ml: Optional[float]
    declared_floz: Optional[float]
    calculated_floz: Optional[float]
    within_tolerance: Optional[bool]
    status_code: StatusCode
    notes: str = ""


@dataclass
class FontMeasurement:
    """Font size measurement from artwork (Section 3E)."""
    text: str
    page_number: int
    font_name: str
    font_size_pt: float
    bbox: Optional[Tuple[float, float, float, float]] = None


@dataclass
class BarcodeResult:
    """Result of barcode scanning and validation (Section 3F)."""
    symbology: str
    decoded_digits: Optional[str]
    printed_digits: Optional[str]
    digits_match: bool
    check_digit_valid: bool
    scan_successful: bool
    quality_notes: str
    status_code: StatusCode


@dataclass
class SnapshotAnnotation:
    """Annotation to be drawn on a snapshot image."""
    issue_id: str
    bbox: Tuple[float, float, float, float]
    status: StatusCode
    page_number: int


@dataclass
class ExtractionResult:
    """Complete extraction result from an artwork file."""
    file_path: str
    file_type: str
    extraction_method: ExtractionMethod
    confidence: float
    text_runs: List[TextRun] = field(default_factory=list)
    pages_processed: int = 0
    pages_with_text: int = 0
    warnings: List[str] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)


@dataclass
class CopyDocument:
    """Parsed copy document data structure."""
    file_path: str
    fields: List[CopyField] = field(default_factory=list)
    instructional_fields: List[CopyField] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)
    project_name: str = ""
    component_type: str = ""


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def fuzzy_ratio(str1: str, str2: str) -> float:
    if not str1 and not str2:
        return 100.0
    if not str1 or not str2:
        return 0.0
    return SequenceMatcher(None, str1, str2).ratio() * 100


def sanitize_for_markdown(text: str) -> str:
    if not text:
        return ""
    text = str(text)
    text = text.replace('\n', ' ').replace('\r', '').replace('\t', ' ')
    text = text.replace('|', '\\|')
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def truncate_text(text: str, max_length: int = 100, suffix: str = "...") -> str:
    if not text or len(text) <= max_length:
        return text or ""
    truncated = text[:max_length - len(suffix)]
    last_space = truncated.rfind(' ')
    if last_space > max_length * 0.5:
        truncated = truncated[:last_space]
    return truncated + suffix


# =============================================================================
# TEXT NORMALIZATION
# =============================================================================

class TextNormalizer:
    """Normalize text for comparison while preserving meaningful differences."""

    @staticmethod
    def normalize(text: str) -> str:
        if not text:
            return ""
        normalized = text
        normalized = re.sub(r'\s+', ' ', normalized).strip()
        # Normalize quotes
        normalized = normalized.replace('\u201C', '"').replace('\u201D', '"')
        normalized = normalized.replace('\u2018', "'").replace('\u2019', "'")
        normalized = normalized.replace('\u00AB', '"').replace('\u00BB', '"')
        # Normalize dashes
        normalized = normalized.replace('\u2013', '-')
        normalized = normalized.replace('\u2010', '-')
        # Normalize apostrophes
        normalized = normalized.replace('\u02BC', "'")
        # Normalize non-breaking spaces
        normalized = normalized.replace('\u00A0', ' ').replace('\u202F', ' ')
        # Normalize ligatures
        normalized = normalized.replace('\uFB01', 'fi').replace('\uFB02', 'fl')
        normalized = normalized.replace('\uFB00', 'ff').replace('\uFB03', 'ffi').replace('\uFB04', 'ffl')
        return normalized

    @staticmethod
    def normalize_for_search(text: str) -> str:
        if not text:
            return ""
        normalized = TextNormalizer.normalize(text).lower()
        normalized = re.sub(r'[^\w\s]', '', normalized)
        normalized = re.sub(r'\s+', ' ', normalized)
        return normalized.strip()


# =============================================================================
# ZOOM TRIGGER DETECTION
# =============================================================================

class ZoomTriggerDetector:
    """Determines when visual verification (zoom) is required."""

    @staticmethod
    def check_triggers(
        text: str,
        font_size: Optional[float] = None,
        confidence: float = 100.0,
        fuzzy_score: float = 100.0
    ) -> Tuple[bool, List[str]]:
        triggers = []

        if font_size is not None and font_size <= config.ZOOM_FONT_SIZE_THRESHOLD:
            triggers.append(f"Font size {font_size}pt <= {config.ZOOM_FONT_SIZE_THRESHOLD}pt threshold")

        if confidence < config.ZOOM_CONFIDENCE_THRESHOLD:
            triggers.append(f"Extraction confidence {confidence:.0f}% < {config.ZOOM_CONFIDENCE_THRESHOLD}%")

        if fuzzy_score < config.ZOOM_FUZZY_THRESHOLD:
            triggers.append(f"Fuzzy match {fuzzy_score:.1f}% < {config.ZOOM_FUZZY_THRESHOLD}%")

        if text:
            if config.ZOOM_ON_NUMBERS and re.search(r'\d', text):
                triggers.append("Contains numbers")
            if config.ZOOM_ON_PERCENTAGE and '%' in text:
                triggers.append("Contains percentage")
            if config.ZOOM_ON_DECIMALS and re.search(r'\d+\.\d+', text):
                triggers.append("Contains decimal numbers")
            if config.ZOOM_ON_UNITS:
                for pattern in config.UNIT_PATTERNS:
                    if re.search(pattern, text, re.IGNORECASE):
                        triggers.append("Contains units")
                        break
            if config.ZOOM_ON_NEGATION:
                text_lower = text.lower()
                for negation in config.NEGATION_WORDS:
                    if re.search(rf'\b{re.escape(negation)}\b', text_lower):
                        triggers.append(f"Contains negation word: '{negation}'")
                        break

        return len(triggers) > 0, triggers


# =============================================================================
# EXCLUSION PATTERN CHECKER
# =============================================================================

class ExclusionChecker:
    """Checks if text should be excluded from analysis (non-product text)."""

    @staticmethod
    def should_exclude(text: str) -> Tuple[bool, Optional[str]]:
        if not text or len(text.strip()) < 2:
            return True, "Empty or too short"

        text_clean = text.strip()
        text_lower = text_clean.lower()

        if text_lower in config.EXCLUSION_EXACT_MATCHES:
            return True, f"Exact match exclusion: '{text_clean}'"

        for company in config.EXCLUSION_COMPANY_NAMES:
            if company.lower() in text_lower:
                return True, f"Company name: '{company}'"

        for pattern in config.EXCLUSION_REGEX_PATTERNS:
            if re.search(pattern, text_clean, re.IGNORECASE):
                return True, f"Pattern match: '{pattern[:30]}...'"

        return False, None

    @staticmethod
    def filter_text_runs(text_runs: List[TextRun]) -> List[TextRun]:
        filtered = []
        excluded_count = 0
        for run in text_runs:
            should_exclude, reason = ExclusionChecker.should_exclude(run.text)
            if not should_exclude:
                filtered.append(run)
            else:
                excluded_count += 1
                logger.debug(f"Excluded text: '{run.text[:50]}' - {reason}")
        if excluded_count > 0:
            logger.info(f"Excluded {excluded_count} non-product text elements")
        return filtered


# =============================================================================
# INSTRUCTIONAL NOTE DETECTOR
# =============================================================================

class InstructionalNoteDetector:
    """Detects internal instructions vs actual copy text."""

    @staticmethod
    def is_instructional(text: str) -> Tuple[bool, Optional[str]]:
        if not text:
            return False, None

        text_stripped = text.strip()

        # Check for empty/placeholder values
        if text_stripped in ['', '-', 'N/A', 'n/a', 'NA', 'na', '\u2014', '\u2013']:
            return True, "Empty/placeholder value"

        for pattern in config.INSTRUCTIONAL_PATTERNS:
            if re.search(pattern, text_stripped, re.IGNORECASE):
                return True, pattern

        return False, None


# =============================================================================
# CURVED TEXT DETECTOR
# =============================================================================

class CurvedTextDetector:
    """Detects garbled text from curved/circular path extraction."""

    @staticmethod
    def is_likely_curved(text: str) -> bool:
        if not text or len(text) < 10:
            return False
        # Heuristic: lots of concatenated words without spaces
        words = text.split()
        if not words:
            return False
        avg_word_len = sum(len(w) for w in words) / len(words)
        # If average word length is very high, likely concatenated
        if avg_word_len > 20:
            return True
        # Check for known garbling patterns
        if re.search(r'[a-z]{3,}[A-Z][a-z]', text):  # camelCase-like from garbling
            return True
        return False


# =============================================================================
# EXCEL COPY DOCUMENT EXTRACTOR
# =============================================================================

class ExcelCopyExtractor:
    """
    Extracts and parses copy documents from .xlsx files.

    Handles the fixed 30-row template structure:
    - Row 1: Headers (language names)
    - Rows 2-5: Front Panel (Product Name, Secondary Name, Fill Weight, Scent)
    - Rows 6-30: Back Panel (Marketing Copy, Claims, Ingredients, etc.)
    - Columns C-M: Language translations
    """

    def __init__(self):
        self.warnings = []

    def extract(self, file_path: Path) -> CopyDocument:
        self.warnings = []

        if not OPENPYXL_AVAILABLE:
            self.warnings.append("openpyxl not available - cannot parse .xlsx files")
            return CopyDocument(file_path=str(file_path), warnings=self.warnings)

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            logger.info(f"Opened Excel copy document: {file_path.name}")

            all_fields = []
            instructional_fields = []
            languages_found = set()

            # Read header row to confirm language columns
            header_langs = {}
            for col_idx, expected_lang in config.LANGUAGE_COLUMNS.items():
                cell_val = ws.cell(row=1, column=col_idx + 1).value  # openpyxl is 1-indexed
                if cell_val:
                    header_langs[col_idx] = cell_val.strip()

            # Process rows 2-30
            for row_num in range(2, 31):
                field_name_cell = ws.cell(row=row_num, column=2)  # Column B
                field_name = str(field_name_cell.value).strip() if field_name_cell.value else ""

                if not field_name:
                    continue

                # Determine panel
                if row_num in config.FRONT_PANEL_ROWS:
                    panel = "Front Panel"
                else:
                    panel = "Back Panel"

                # Override display name for Hero Ingredients Header (row 16)
                display_name = None
                if row_num == config.HERO_INGREDIENTS_HEADER_ROW:
                    display_name = "Hero Ingredients Header"

                # Process each language column
                for col_idx, lang_name in config.LANGUAGE_COLUMNS.items():
                    cell = ws.cell(row=row_num, column=col_idx + 1)
                    cell_value = cell.value

                    if cell_value is None or str(cell_value).strip() == "":
                        continue  # Skip empty cells

                    text = str(cell_value).strip()

                    # Check for underline formatting
                    is_underlined = False
                    if cell.font and cell.font.underline:
                        is_underlined = True

                    # Check if this is an instructional note
                    is_instr, instr_pattern = InstructionalNoteDetector.is_instructional(text)

                    languages_found.add(lang_name)

                    copy_field = CopyField(
                        field_name=field_name,
                        panel=panel,
                        language=lang_name,
                        text=text,
                        row_number=row_num,
                        col_index=col_idx,
                        is_instructional=is_instr,
                        instructional_pattern=instr_pattern,
                        is_underlined=is_underlined,
                        display_name=display_name,
                    )

                    if is_instr:
                        instructional_fields.append(copy_field)
                    else:
                        all_fields.append(copy_field)

            # Build project name (before closing workbook)
            product_name = self._get_cell_value(ws, 2, 3)  # C2
            secondary_name = self._get_cell_value(ws, 3, 3)  # C3
            fill_weight_raw = self._get_cell_value(ws, 4, 3)  # C4
            metric_volume = self._extract_metric_volume(fill_weight_raw)
            project_name = self._build_project_name(product_name, secondary_name, metric_volume)

            # Extract component type from filename
            component_type = self._extract_component_type(file_path)

            # Check for translation existence issues
            translation_warnings = self._check_translation_existence(ws)

            wb.close()

            copy_doc = CopyDocument(
                file_path=str(file_path),
                fields=all_fields,
                instructional_fields=instructional_fields,
                warnings=self.warnings + translation_warnings,
                metadata={
                    'total_fields': len(all_fields),
                    'instructional_count': len(instructional_fields),
                    'languages_found': list(languages_found),
                    'header_languages': header_langs,
                },
                project_name=project_name,
                component_type=component_type,
            )

            logger.info(f"Extracted {len(all_fields)} active fields, "
                        f"{len(instructional_fields)} instructional notes")
            return copy_doc

        except Exception as e:
            self.warnings.append(f"Extraction error: {str(e)}")
            logger.error(f"Failed to extract Excel copy document: {e}")
            return CopyDocument(file_path=str(file_path), warnings=self.warnings)

    def _get_cell_value(self, ws, row: int, col: int) -> str:
        """Safely get a cell value as string."""
        # Re-open workbook if ws is closed; for safety we accept ws directly
        try:
            val = ws.cell(row=row, column=col).value
            return str(val).strip() if val else ""
        except:
            return ""

    def _extract_metric_volume(self, fill_weight: str) -> str:
        if not fill_weight:
            return ""
        match = re.search(r'(\d+(?:\.\d+)?)\s*(?:ML|ml|mL)', fill_weight, re.IGNORECASE)
        if match:
            return f"{match.group(1)} ml"
        return ""

    def _build_project_name(self, product_name: str, secondary_name: str, volume: str) -> str:
        parts = [config.BRAND_NAME]
        if product_name:
            parts.append(product_name)
        if secondary_name:
            parts.append(secondary_name)
        if volume:
            parts.append(volume)
        return " ".join(parts)

    def _extract_component_type(self, file_path: Path) -> str:
        stem = file_path.stem  # filename without extension
        # Extract text after last "_-_"
        if '_-_' in stem:
            component = stem.rsplit('_-_', 1)[-1]
            return component.replace('_', ' ').strip()
        # Fallback: try last segment after "__"
        if '__' in stem:
            component = stem.rsplit('__', 1)[-1]
            return component.replace('_', ' ').strip()
        return stem

    def _check_translation_existence(self, ws) -> List[str]:
        """Check if non-English columns have content but English is empty."""
        warnings = []
        for row_num in range(2, 31):
            field_name = ws.cell(row=row_num, column=2).value
            if not field_name:
                continue

            english_val = ws.cell(row=row_num, column=3).value  # Column C
            english_empty = english_val is None or str(english_val).strip() == ""

            for col_idx, lang_name in config.LANGUAGE_COLUMNS.items():
                if lang_name == 'English':
                    continue
                other_val = ws.cell(row=row_num, column=col_idx + 1).value
                if other_val and str(other_val).strip() and english_empty:
                    warnings.append(
                        f"Row {row_num} ({field_name}): {lang_name} has content but English is empty"
                    )
        return warnings


# =============================================================================
# INCI CAPITALIZATION VALIDATOR
# =============================================================================

class INCIValidator:
    """Validates INCI capitalization rules."""

    @staticmethod
    def validate_inci_capitalization(text: str) -> Tuple[bool, List[str]]:
        if not text:
            return True, []
        issues = []
        words = re.findall(r'\(?\b\w+\b\)?', text)
        for word in words:
            clean_word = word.strip('()')
            if clean_word.lower() in config.INCI_LOWERCASE_CONNECTORS:
                if clean_word[0].isupper() and clean_word.lower() in [
                    'de', 'du', 'des', 'la', 'le', 'les', 'of', 'the', 'and', 'with', 'et', 'cum'
                ]:
                    issues.append(f"'{clean_word}' should be lowercase '{clean_word.lower()}'")
                continue
            if len(clean_word) < 2 or clean_word.isdigit():
                continue
            if clean_word.isupper():
                continue
            if clean_word.upper() in config.ALLOWED_ACRONYMS:
                continue
            if clean_word[0].islower() and clean_word[0].isalpha():
                correct = clean_word[0].upper() + clean_word[1:]
                issues.append(f"'{clean_word}' should be '{correct}'")
        return len(issues) == 0, issues

    @staticmethod
    def validate_ingredient_list(ingredient_text: str) -> List[Dict[str, str]]:
        findings = []
        ingredients = re.split(r'[,;]', ingredient_text)
        for ingredient in ingredients:
            ingredient = ingredient.strip()
            if not ingredient:
                continue
            is_valid, issues = INCIValidator.validate_inci_capitalization(ingredient)
            if not is_valid:
                findings.append({
                    'ingredient': ingredient[:50],
                    'issues': issues,
                    'severity': 'ATTN'
                })
        return findings


# =============================================================================
# COPY QUALITY CHECKER (Section 3A)
# =============================================================================

class CopyQualityChecker:
    """
    Checks copy document text for quality issues.
    
    Rules:
    - Rows 2, 3, 5-22: must be lowercase across all columns
    - Exception: 2+ consecutive uppercase letters = acceptable (acronyms, language prefixes)
    - Row 4 (Fill Weight): exempt from case checking
    - Rows 23-27: exempt from lowercase checking (Address, Biorius, Country, Ingredients, Social)
    - Standard punctuation checks
    - Instructional note detection
    - Translation existence check
    """

    @staticmethod
    def _has_consecutive_uppercase(word: str, min_count: int = 2) -> bool:
        """Check if a word has 2+ consecutive uppercase letters."""
        count = 0
        for char in word:
            if char.isupper():
                count += 1
                if count >= min_count:
                    return True
            else:
                count = 0
        return False

    @staticmethod
    def check_capitalization(
        text: str,
        field_name: str,
        language: str,
        row_number: int
    ) -> List[CopyQualityIssue]:
        issues = []
        if not text:
            return issues

        # Skip rows not subject to lowercase checking
        if row_number not in config.LOWERCASE_CHECK_ROWS:
            return issues

        # Skip fields explicitly exempt
        for allowed in config.UPPERCASE_ALLOWED_FIELDS:
            if allowed.lower() in field_name.lower():
                return issues

        # Find words with uppercase that shouldn't be
        words = re.findall(r'\b[A-Za-z][A-Za-z\']+\b', text)

        for word in words:
            # If 2+ consecutive uppercase letters, acceptable (acronym/prefix)
            if CopyQualityChecker._has_consecutive_uppercase(word):
                continue

            # Single uppercase letter at start (not after sentence-ending punctuation)
            if word[0].isupper():
                word_pos = text.find(word)
                before = text[:word_pos].rstrip()
                # Acceptable if after sentence-ending punctuation
                if before and before[-1] in '.!?':
                    continue
                # Acceptable if it's the very first word in the cell
                if word_pos == 0:
                    # Still flag - copy should be all lowercase
                    pass

                issues.append(CopyQualityIssue(
                    language=language,
                    field_name=field_name,
                    original_text=word,
                    issue_type="Capitalization",
                    recommendation=f"'{word}' should be lowercase '{word.lower()}'",
                    status_code=StatusCode.ATTN
                ))

        return issues

    @staticmethod
    def check_punctuation(
        text: str,
        field_name: str,
        language: str
    ) -> List[CopyQualityIssue]:
        issues = []
        if not text:
            return issues

        if re.search(r'[.]{2,}(?!\.)', text):
            issues.append(CopyQualityIssue(
                language=language, field_name=field_name,
                original_text=text[:50], issue_type="Punctuation",
                recommendation="Check for double periods",
                status_code=StatusCode.ATTN
            ))

        if ',,' in text:
            issues.append(CopyQualityIssue(
                language=language, field_name=field_name,
                original_text=text[:50], issue_type="Punctuation",
                recommendation="Remove double commas",
                status_code=StatusCode.ATTN
            ))

        if '  ' in text:
            issues.append(CopyQualityIssue(
                language=language, field_name=field_name,
                original_text=text[:50], issue_type="Formatting",
                recommendation="Remove extra spaces",
                status_code=StatusCode.ATTN
            ))

        return issues

    @staticmethod
    def check_translation_quality(
        copy_doc: CopyDocument
    ) -> List[CopyQualityIssue]:
        """Check that translations exist where English has content (and vice versa)."""
        issues = []
        # Group fields by row
        fields_by_row = {}
        for f in copy_doc.fields:
            if f.row_number not in fields_by_row:
                fields_by_row[f.row_number] = {}
            fields_by_row[f.row_number][f.language] = f

        # Also include instructional fields for completeness
        for f in copy_doc.instructional_fields:
            if f.row_number not in fields_by_row:
                fields_by_row[f.row_number] = {}
            fields_by_row[f.row_number][f.language] = f

        # Check: non-English has content but English is empty
        for row_num, lang_fields in fields_by_row.items():
            has_english = 'English' in lang_fields
            for lang, f in lang_fields.items():
                if lang == 'English':
                    continue
                if not has_english and f.text.strip():
                    issues.append(CopyQualityIssue(
                        language=lang,
                        field_name=f.field_name,
                        original_text=f.text[:50],
                        issue_type="Missing English",
                        recommendation=f"{lang} has content but English is empty for this row",
                        status_code=StatusCode.ATTN
                    ))

        return issues

    @classmethod
    def analyze_copy_document(cls, copy_doc: CopyDocument) -> List[CopyQualityIssue]:
        all_issues = []

        # Check active fields
        for f in copy_doc.fields:
            all_issues.extend(cls.check_capitalization(
                f.text, f.field_name, f.language, f.row_number
            ))
            all_issues.extend(cls.check_punctuation(
                f.text, f.field_name, f.language
            ))

        # Flag instructional notes (as ATTN, not FAIL)
        for f in copy_doc.instructional_fields:
            all_issues.append(CopyQualityIssue(
                language=f.language,
                field_name=f.field_name,
                original_text=truncate_text(f.text, 60),
                issue_type="Instructional note",
                recommendation=f"Remove internal instruction (matched: {truncate_text(f.instructional_pattern or 'N/A', 30)})",
                status_code=StatusCode.ATTN
            ))

        # Check INCI in ingredient fields
        for f in copy_doc.fields:
            if 'ingredient list' in f.field_name.lower():
                # Strip formula number for INCI check
                inci_text = re.sub(r'^\(\d+\)\s*', '', f.text)
                inci_findings = INCIValidator.validate_ingredient_list(inci_text)
                for finding in inci_findings:
                    all_issues.append(CopyQualityIssue(
                        language=f.language,
                        field_name=f.field_name,
                        original_text=finding['ingredient'],
                        issue_type="INCI Capitalization",
                        recommendation="; ".join(finding['issues']),
                        status_code=StatusCode.ATTN
                    ))

        # Translation existence checks
        all_issues.extend(cls.check_translation_quality(copy_doc))

        logger.info(f"Copy quality analysis found {len(all_issues)} issues")
        return all_issues


# =============================================================================
# CLAIM RISK ASSESSOR (Section 3B)
# =============================================================================

class ClaimRiskAssessor:
    """Assesses regulatory risk of product claims. Only checks Pack Claims (rows 11-15)."""

    @staticmethod
    def assess_claim(claim_text: str, language: str) -> ClaimRisk:
        claim_lower = claim_text.lower()

        for term in config.HIGH_RISK_CLAIM_TERMS:
            if term.lower() in claim_lower:
                return ClaimRisk(
                    language=language, claim_text=claim_text,
                    risk_level=RiskLevel.HIGH,
                    rationale=f"Contains high-risk term: '{term}' - requires substantiation",
                    regions=["USA", "EU", "UK", "CA"],
                    recommended_action="Escalate",
                    status_code=StatusCode.ATTN
                )

        for term in config.MEDIUM_RISK_CLAIM_TERMS:
            if term.lower() in claim_lower:
                return ClaimRisk(
                    language=language, claim_text=claim_text,
                    risk_level=RiskLevel.MEDIUM,
                    rationale=f"Contains efficacy claim: '{term}' - verify support",
                    regions=["USA", "EU", "UK"],
                    recommended_action="Verify",
                    status_code=StatusCode.ATTN
                )

        return ClaimRisk(
            language=language, claim_text=claim_text,
            risk_level=RiskLevel.LOW,
            rationale="Cosmetic/descriptive claim - acceptable",
            regions=["All"],
            recommended_action="Keep",
            status_code=StatusCode.OK
        )

    @classmethod
    def assess_all_claims(cls, copy_doc: CopyDocument) -> List[ClaimRisk]:
        assessments = []
        for f in copy_doc.fields:
            # Only check Pack Claims (rows 11-15), English only
            if f.row_number not in config.PACK_CLAIM_ROWS:
                continue
            if f.language != 'English':
                continue
            if f.is_instructional:
                continue
            assessment = cls.assess_claim(f.text, f.language)
            assessments.append(assessment)
        logger.info(f"Assessed {len(assessments)} claims")
        return assessments


# =============================================================================
# CONVERSION CHECKER (Section 3C)
# =============================================================================

class ConversionChecker:
    """Verifies volume conversions between mL and fl oz from Fill Weight (row 4)."""

    ML_TO_FL_OZ = 0.033814
    TOLERANCE = 0.10

    @staticmethod
    def check_conversions(copy_doc: CopyDocument) -> List[ConversionCheck]:
        results = []

        # Find fill weight field (row 4, English)
        fill_weight_field = None
        for f in copy_doc.fields:
            if f.row_number == config.FILL_WEIGHT_ROW and f.language == 'English':
                fill_weight_field = f
                break

        # Check if fill weight was flagged as instructional
        if fill_weight_field is None:
            for f in copy_doc.instructional_fields:
                if f.row_number == config.FILL_WEIGHT_ROW and f.language == 'English':
                    results.append(ConversionCheck(
                        source_field="Fill Weight",
                        declared_ml=None, declared_floz=None,
                        calculated_floz=None, within_tolerance=None,
                        status_code=StatusCode.ATTN,
                        notes=f"Fill weight marked as instructional — {f.text}"
                    ))
                    return results

            # No fill weight at all
            logger.warning("No Fill Weight field found")
            return results

        # Extract volumes
        ml_value, fl_oz_value = ConversionChecker._extract_volumes(fill_weight_field.text)

        if ml_value and fl_oz_value:
            calculated = ml_value * ConversionChecker.ML_TO_FL_OZ
            difference = abs(fl_oz_value - calculated)
            passed = difference <= ConversionChecker.TOLERANCE

            results.append(ConversionCheck(
                source_field="Fill Weight",
                declared_ml=ml_value,
                declared_floz=fl_oz_value,
                calculated_floz=round(calculated, 2),
                within_tolerance=passed,
                status_code=StatusCode.OK if passed else StatusCode.FAIL,
                notes=f"Declared: {fl_oz_value} fl oz, Calculated: {calculated:.2f} fl oz, Diff: {difference:.2f}"
            ))
        else:
            logger.warning(f"Could not extract volumes from: {fill_weight_field.text}")

        return results

    @staticmethod
    def _extract_volumes(text: str) -> Tuple[Optional[float], Optional[float]]:
        ml_pattern = r'(\d+(?:\.\d+)?)\s*(?:ML|ml|mL|Ml)\b'
        fl_oz_pattern = r'(\d+(?:\.\d+)?)\s*(?:US\s*)?(?:FL\.?\s*OZ\.?|fl\.?\s*oz\.?)\b'
        ml_match = re.search(ml_pattern, text, re.IGNORECASE)
        fl_oz_match = re.search(fl_oz_pattern, text, re.IGNORECASE)
        ml_value = float(ml_match.group(1)) if ml_match else None
        fl_oz_value = float(fl_oz_match.group(1)) if fl_oz_match else None
        return ml_value, fl_oz_value


# =============================================================================
# PDF EXTRACTOR
# =============================================================================

class PDFExtractor:
    """Extracts text from PDF artwork files using PyMuPDF."""

    def __init__(self):
        self.warnings = []

    def extract(self, file_path: Path) -> ExtractionResult:
        self.warnings = []
        if not PYMUPDF_AVAILABLE:
            self.warnings.append("PyMuPDF (fitz) not available")
            return ExtractionResult(
                file_path=str(file_path), file_type="pdf",
                extraction_method=ExtractionMethod.FAILED,
                confidence=0.0, warnings=self.warnings
            )

        try:
            doc = fitz.open(file_path)
            page_count = len(doc)
            logger.info(f"Opened PDF: {file_path.name} ({page_count} pages)")

            all_text_runs = []
            pages_with_text = 0

            for page_num in range(page_count):
                page = doc[page_num]
                page_runs = self._extract_page_text(page, page_num + 1)
                if page_runs:
                    all_text_runs.extend(page_runs)
                    pages_with_text += 1
                else:
                    self.warnings.append(f"Page {page_num + 1}: No extractable text found")

            if page_count == 0:
                confidence = 0.0
                method = ExtractionMethod.FAILED
            elif pages_with_text == 0:
                confidence = 0.0
                method = ExtractionMethod.FAILED
                self.warnings.append("No text extracted - PDF may have outlined fonts")
            else:
                confidence = pages_with_text / page_count
                method = ExtractionMethod.LIVE_TEXT

            doc.close()
            filtered_runs = ExclusionChecker.filter_text_runs(all_text_runs)

            return ExtractionResult(
                file_path=str(file_path), file_type="pdf",
                extraction_method=method, confidence=confidence,
                text_runs=filtered_runs,
                pages_processed=page_count,
                pages_with_text=pages_with_text,
                warnings=self.warnings,
                metadata={
                    'total_runs_before_filter': len(all_text_runs),
                    'total_runs_after_filter': len(filtered_runs),
                    'page_count': page_count,
                }
            )

        except Exception as e:
            self.warnings.append(f"PDF extraction error: {str(e)}")
            logger.error(f"Failed to extract PDF: {e}")
            return ExtractionResult(
                file_path=str(file_path), file_type="pdf",
                extraction_method=ExtractionMethod.FAILED,
                confidence=0.0, warnings=self.warnings
            )

    def _extract_page_text(self, page, page_num: int) -> List[TextRun]:
        text_runs = []
        try:
            blocks = page.get_text("dict", flags=11)["blocks"]
            for block in blocks:
                if block.get("type") != 0:
                    continue
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        text = span.get("text", "").strip()
                        if not text:
                            continue
                        bbox = span.get("bbox")
                        font_name = span.get("font", "Unknown")
                        font_size = span.get("size", 0.0)
                        rotation = self._detect_rotation(line)

                        # Detect underline from flags
                        flags = span.get("flags", 0)
                        is_underlined = bool(flags & 4)  # Bit 2 = underline

                        text_runs.append(TextRun(
                            text=text, page_number=page_num,
                            bbox=tuple(bbox) if bbox else None,
                            font_name=font_name,
                            font_size=round(font_size, 2),
                            extraction_method=ExtractionMethod.LIVE_TEXT,
                            confidence=1.0, rotation=rotation,
                            is_underlined=is_underlined,
                        ))

            # Fallback: simple text extraction
            if not text_runs:
                simple_text = page.get_text().strip()
                if simple_text and len(simple_text) >= 10:
                    text_runs.append(TextRun(
                        text=simple_text, page_number=page_num,
                        extraction_method=ExtractionMethod.LIVE_TEXT,
                        confidence=0.7
                    ))

            # Reconstruct rotated text
            reconstructed = self._reconstruct_rotated_runs(text_runs, page_num)
            if reconstructed:
                logger.debug(f"Page {page_num}: added {len(reconstructed)} reconstructed rotated run(s)")
                text_runs.extend(reconstructed)

        except Exception as e:
            logger.warning(f"Page {page_num} extraction error: {e}")
        return text_runs

    def _detect_rotation(self, line: Dict) -> int:
        dir_vec = line.get("dir", (1, 0))
        if len(dir_vec) < 2:
            return 0
        dx, dy = dir_vec
        if abs(dx - 1) < 0.1 and abs(dy) < 0.1:
            return 0
        elif abs(dx) < 0.1 and abs(dy - 1) < 0.1:
            return 90
        elif abs(dx + 1) < 0.1 and abs(dy) < 0.1:
            return 180
        elif abs(dx) < 0.1 and abs(dy + 1) < 0.1:
            return 270
        return 0

    def _reconstruct_rotated_runs(self, text_runs: List[TextRun], page_num: int) -> List[TextRun]:
        reconstructed = []
        for rotation in (90, 180, 270):
            rotated = [r for r in text_runs if r.rotation == rotation and r.bbox]
            if len(rotated) < 2:
                continue
            groups = self._group_rotated_by_column(rotated, rotation)
            for group in groups:
                if len(group) < 2:
                    continue
                if rotation == 90:
                    group.sort(key=lambda r: -(r.bbox[1] + r.bbox[3]) / 2)
                elif rotation == 270:
                    group.sort(key=lambda r: (r.bbox[1] + r.bbox[3]) / 2)
                elif rotation == 180:
                    group.sort(key=lambda r: -(r.bbox[0] + r.bbox[2]) / 2)

                combined = ' '.join(r.text for r in group if r.text.strip())
                if not combined.strip():
                    continue
                x0 = min(r.bbox[0] for r in group)
                y0 = min(r.bbox[1] for r in group)
                x1 = max(r.bbox[2] for r in group)
                y1 = max(r.bbox[3] for r in group)
                reconstructed.append(TextRun(
                    text=combined, page_number=page_num,
                    bbox=(x0, y0, x1, y1),
                    font_name=group[0].font_name,
                    font_size=group[0].font_size,
                    extraction_method=ExtractionMethod.LIVE_TEXT,
                    confidence=1.0, rotation=rotation
                ))
        return reconstructed

    def _group_rotated_by_column(self, runs: List[TextRun], rotation: int) -> List[List[TextRun]]:
        threshold = config.ROTATION_PROXIMITY_THRESHOLD
        def axis_center(run):
            if rotation in (90, 270):
                return (run.bbox[0] + run.bbox[2]) / 2
            return (run.bbox[1] + run.bbox[3]) / 2

        sorted_runs = sorted(runs, key=axis_center)
        groups = []
        current = [sorted_runs[0]]
        current_key = axis_center(sorted_runs[0])
        for run in sorted_runs[1:]:
            key = axis_center(run)
            if abs(key - current_key) <= threshold:
                current.append(run)
            else:
                groups.append(current)
                current = [run]
                current_key = key
        groups.append(current)
        return groups


# =============================================================================
# AI FILE EXTRACTOR
# =============================================================================

class AIExtractor:
    """Extracts text from .ai files via PDF compatibility."""

    def __init__(self):
        self.pdf_extractor = PDFExtractor()

    def extract(self, file_path: Path) -> ExtractionResult:
        result = self.pdf_extractor.extract(file_path)
        if result.confidence > 0:
            result.extraction_method = ExtractionMethod.AI_PDF_COMPATIBLE
            result.file_type = "ai"
            result.warnings.insert(0, "AI file extracted via PDF compatibility")
        else:
            result.warnings.append("AI file may not have PDF compatibility enabled")
        return result


# =============================================================================
# FONT EXTRACTOR (Section 3E)
# =============================================================================

class FontExtractor:
    """Extracts font size metadata from artwork."""

    def extract_from_pdf(self, file_path: Path) -> List[FontMeasurement]:
        if not PYMUPDF_AVAILABLE:
            return []
        measurements = []
        try:
            doc = fitz.open(file_path)
            for page_num in range(len(doc)):
                page = doc[page_num]
                blocks = page.get_text("dict", flags=11)["blocks"]
                for block in blocks:
                    if block.get("type") != 0:
                        continue
                    for line in block.get("lines", []):
                        for span in line.get("spans", []):
                            text = span.get("text", "").strip()
                            if not text:
                                continue
                            should_exclude, _ = ExclusionChecker.should_exclude(text)
                            if should_exclude:
                                continue
                            font_size = span.get("size", 0)
                            if font_size > 0:
                                measurements.append(FontMeasurement(
                                    text=text[:50], page_number=page_num + 1,
                                    font_name=span.get("font", "Unknown"),
                                    font_size_pt=round(font_size, 2),
                                    bbox=tuple(span.get("bbox", []))
                                ))
            doc.close()
        except Exception as e:
            logger.error(f"Font extraction error: {e}")
        return measurements

    def get_smallest_font(self, measurements: List[FontMeasurement]) -> Optional[FontMeasurement]:
        if not measurements:
            return None
        return min(measurements, key=lambda m: m.font_size_pt)


# =============================================================================
# BARCODE SCANNER (Section 3F)
# =============================================================================

class BarcodeScanner:
    """Scans and validates barcodes in artwork."""

    @staticmethod
    def scan_from_pdf(file_path: Path) -> List[BarcodeResult]:
        results = []
        if not PYMUPDF_AVAILABLE or not PIL_AVAILABLE:
            return [BarcodeResult(
                symbology="Unknown", decoded_digits=None, printed_digits=None,
                digits_match=False, check_digit_valid=False, scan_successful=False,
                quality_notes="Required libraries not available",
                status_code=StatusCode.TBD
            )]
        if not PYZBAR_AVAILABLE:
            return [BarcodeResult(
                symbology="Unknown", decoded_digits=None, printed_digits=None,
                digits_match=False, check_digit_valid=False, scan_successful=False,
                quality_notes="pyzbar not available - manual verification required",
                status_code=StatusCode.TBD
            )]
        try:
            doc = fitz.open(file_path)
            for page_num in range(len(doc)):
                page = doc[page_num]
                mat = fitz.Matrix(config.BARCODE_RENDER_DPI, config.BARCODE_RENDER_DPI)
                pix = page.get_pixmap(matrix=mat)
                img_data = pix.tobytes("png")
                img = Image.open(BytesIO(img_data))
                decoded = decode_barcode(img)
                for barcode in decoded:
                    barcode_type = barcode.type
                    barcode_data = barcode.data.decode('utf-8')
                    is_valid = BarcodeScanner._validate_check_digit(barcode_data, barcode_type)
                    results.append(BarcodeResult(
                        symbology=barcode_type, decoded_digits=barcode_data,
                        printed_digits=barcode_data,
                        digits_match=True, check_digit_valid=is_valid,
                        scan_successful=True,
                        quality_notes=f"Decoded on page {page_num + 1}",
                        status_code=StatusCode.OK if is_valid else StatusCode.FAIL
                    ))
            doc.close()
            if not results:
                results.append(BarcodeResult(
                    symbology="N/A", decoded_digits=None, printed_digits=None,
                    digits_match=False, check_digit_valid=False, scan_successful=False,
                    quality_notes="No barcode detected in artwork",
                    status_code=StatusCode.FYI
                ))
        except Exception as e:
            logger.error(f"Barcode scanning error: {e}")
            results.append(BarcodeResult(
                symbology="Unknown", decoded_digits=None, printed_digits=None,
                digits_match=False, check_digit_valid=False, scan_successful=False,
                quality_notes=f"Scan error: {str(e)}",
                status_code=StatusCode.TBD
            ))
        return results

    @staticmethod
    def _validate_check_digit(digits: str, barcode_type: str) -> bool:
        digits = digits.replace(" ", "").replace("-", "")
        if barcode_type in ["EAN13", "EAN-13"] and len(digits) == 13:
            return BarcodeScanner._ean13(digits)
        elif barcode_type in ["UPCA", "UPC-A", "UPC_A"] and len(digits) == 12:
            return BarcodeScanner._upc_a(digits)
        elif barcode_type in ["EAN8", "EAN-8"] and len(digits) == 8:
            return BarcodeScanner._ean8(digits)
        return True

    @staticmethod
    def _ean13(d):
        if len(d) != 13 or not d.isdigit(): return False
        o = sum(int(d[i]) for i in range(0, 12, 2))
        e = sum(int(d[i]) for i in range(1, 12, 2))
        return (10 - ((o + e * 3) % 10)) % 10 == int(d[12])

    @staticmethod
    def _upc_a(d):
        if len(d) != 12 or not d.isdigit(): return False
        o = sum(int(d[i]) for i in range(0, 11, 2))
        e = sum(int(d[i]) for i in range(1, 11, 2))
        return (10 - ((o * 3 + e) % 10)) % 10 == int(d[11])

    @staticmethod
    def _ean8(d):
        if len(d) != 8 or not d.isdigit(): return False
        o = sum(int(d[i]) for i in range(0, 7, 2))
        e = sum(int(d[i]) for i in range(1, 7, 2))
        return (10 - ((o * 3 + e) % 10)) % 10 == int(d[7])


# =============================================================================
# ARTWORK MATCHER (Section 3D)
# =============================================================================

class ArtworkMatcher:
    """
    Matches copy document fields against extracted artwork text.
    
    Anti-hallucination invariants:
    1. extraction_method == FAILED -> NO matches generated
    2. No matching text_run found -> MISSING_IN_ARTWORK
    3. EXACT_MATCH only if normalized strings are identical
    4. All scores grounded in actual fuzzy match calculations
    5. Cross-references Section 3A flags in notes
    """

    def __init__(self):
        self.normalizer = TextNormalizer
        self.zoom_detector = ZoomTriggerDetector

    def match_fields(
        self,
        copy_fields: List[CopyField],
        text_runs: List[TextRun],
        extraction_method: ExtractionMethod,
        quality_issues: List[CopyQualityIssue]
    ) -> List[MatchFinding]:
        findings = []

        # Build 3A lookup: (field_name, language) -> list of issues
        quality_lookup = {}
        for issue in quality_issues:
            key = (issue.field_name, issue.language)
            if key not in quality_lookup:
                quality_lookup[key] = []
            quality_lookup[key].append(issue)

        # FAIL-FAST: If extraction failed
        if extraction_method == ExtractionMethod.FAILED:
            logger.warning("Extraction failed - cannot generate reliable matches")
            for f in copy_fields:
                findings.append(MatchFinding(
                    field_name=f.display_name or f.field_name,
                    panel=f.panel, language=f.language,
                    copy_value=f.text, artwork_value=None,
                    match_type=MatchType.REQUIRES_VERIFICATION,
                    similarity_score=0.0,
                    status_code=StatusCode.TBD,
                    notes=["Extraction failed - manual verification required"]
                ))
            return findings

        artwork_lookup = self._build_text_lookup(text_runs)
        issue_counter = 1

        for f in copy_fields:
            # Skip instructional fields (already excluded from copy_doc.fields)
            display_name = f.display_name or f.field_name

            # Handle ingredient list specially
            if 'ingredient list' in f.field_name.lower():
                ingredient_findings = self._match_ingredient_list(
                    f, text_runs, artwork_lookup, quality_lookup, issue_counter
                )
                for ifind in ingredient_findings:
                    findings.append(ifind)
                    if ifind.status_code != StatusCode.OK:
                        issue_counter += 1
                continue

            finding = self._match_single_field(
                f, text_runs, artwork_lookup, quality_lookup
            )

            # Assign issue ID if not OK
            if finding.status_code != StatusCode.OK:
                finding.issue_id = f"D-{issue_counter:03d}"
                issue_counter += 1

            findings.append(finding)

        # Count deferred fields (instructional) for summary
        self._log_match_summary(findings)
        return findings

    def _match_ingredient_list(
        self,
        field: CopyField,
        text_runs: List[TextRun],
        artwork_lookup: Dict,
        quality_lookup: Dict,
        issue_counter: int
    ) -> List[MatchFinding]:
        """Match ingredient list as two separate items: formula number + ingredient text."""
        findings = []
        text = field.text

        # Extract formula number
        formula_match = re.match(r'\((\d+)\)\s*(.*)', text, re.DOTALL)
        if formula_match:
            formula_num = formula_match.group(1)
            ingredient_text = formula_match.group(2).strip()

            # 1. Check formula number in artwork
            formula_found = False
            for run in text_runs:
                if f"({formula_num})" in run.text or formula_num in run.text:
                    formula_found = True
                    break

            formula_finding = MatchFinding(
                field_name="Ingredient List (Formula #)",
                panel=field.panel, language=field.language,
                copy_value=f"({formula_num})",
                artwork_value=f"({formula_num})" if formula_found else None,
                match_type=MatchType.EXACT_MATCH if formula_found else MatchType.MISSING_IN_ARTWORK,
                similarity_score=100.0 if formula_found else 0.0,
                status_code=StatusCode.OK if formula_found else StatusCode.FAIL,
                notes=["Formula number match" if formula_found else "Formula number NOT FOUND in artwork"],
                requires_zoom=True,
                zoom_reasons=["Contains numbers"],
            )
            if not formula_found:
                formula_finding.issue_id = f"D-{issue_counter:03d}"
            findings.append(formula_finding)

            # 2. Check ingredient text
            has_3a = (field.field_name, field.language) in quality_lookup
            ingredient_finding = self._match_text_against_runs(
                "Ingredient List", field.panel, field.language,
                ingredient_text, text_runs, artwork_lookup, has_3a
            )
            if ingredient_finding.status_code != StatusCode.OK:
                # Offset 0 when formula was found (formula ID slot not consumed),
                # offset 1 when formula was NOT found (formula consumed issue_counter).
                ingredient_finding.issue_id = f"D-{issue_counter + (0 if formula_found else 1):03d}"
            findings.append(ingredient_finding)
        else:
            # No formula number, match whole text
            has_3a = (field.field_name, field.language) in quality_lookup
            finding = self._match_text_against_runs(
                "Ingredient List", field.panel, field.language,
                text, text_runs, artwork_lookup, has_3a
            )
            findings.append(finding)

        return findings

    def _match_single_field(
        self,
        field: CopyField,
        text_runs: List[TextRun],
        artwork_lookup: Dict,
        quality_lookup: Dict
    ) -> MatchFinding:
        display_name = field.display_name or field.field_name
        has_3a = (field.field_name, field.language) in quality_lookup

        finding = self._match_text_against_runs(
            display_name, field.panel, field.language,
            field.text, text_runs, artwork_lookup, has_3a
        )

        # Check underline discrepancy
        if finding.matched_runs:
            artwork_underlined = any(r.is_underlined for r in finding.matched_runs)
            if field.is_underlined != artwork_underlined:
                if artwork_underlined and not field.is_underlined:
                    finding.notes.append("Text is underlined in artwork but not in copy document")
                    if finding.status_code == StatusCode.OK:
                        finding.status_code = StatusCode.ATTN
                elif field.is_underlined and not artwork_underlined:
                    finding.notes.append("Text is underlined in copy document but not in artwork")
                    if finding.status_code == StatusCode.OK:
                        finding.status_code = StatusCode.ATTN

        return finding

    def _match_text_against_runs(
        self,
        field_name: str,
        panel: str,
        language: str,
        copy_text: str,
        text_runs: List[TextRun],
        artwork_lookup: Dict,
        has_3a_flag: bool
    ) -> MatchFinding:
        """Core matching logic with anti-hallucination safeguards."""
        normalized_copy = self.normalizer.normalize(copy_text)

        requires_visual, zoom_triggers = self.zoom_detector.check_triggers(copy_text)
        if config.VISION_ALWAYS_REQUIRED and not requires_visual:
            requires_visual = True
            zoom_triggers = ["AI vision double-check always required"] + zoom_triggers

        # 3A cross-reference note
        cross_ref_note = ""
        if has_3a_flag:
            cross_ref_note = " | Note: copy doc flagged in A"

        # Try exact match
        if normalized_copy in artwork_lookup:
            status = StatusCode.ATTN if has_3a_flag else (StatusCode.TBD if requires_visual else StatusCode.OK)
            notes = []
            if has_3a_flag:
                notes.append("Matches copy doc but flagged in A — see copy quality issue")
            else:
                notes.append("Exact match")
            return MatchFinding(
                field_name=field_name, panel=panel, language=language,
                copy_value=copy_text, artwork_value=copy_text,
                match_type=MatchType.EXACT_MATCH,
                similarity_score=100.0,
                status_code=status,
                requires_zoom=requires_visual,
                zoom_reasons=zoom_triggers,
                notes=notes,
                matched_runs=artwork_lookup[normalized_copy],
                has_3a_flag=has_3a_flag,
            )

        # Fuzzy matching with sliding window
        best_match = self._fuzzy_match_sliding_window(normalized_copy, text_runs)

        if best_match:
            match_type, fuzzy_score, matched_text, matched_runs = best_match

            # Check if garbled (curved text)
            if CurvedTextDetector.is_likely_curved(matched_text) and fuzzy_score < 80:
                return MatchFinding(
                    field_name=field_name, panel=panel, language=language,
                    copy_value=copy_text, artwork_value=matched_text,
                    match_type=MatchType.REQUIRES_VERIFICATION,
                    similarity_score=fuzzy_score,
                    status_code=StatusCode.ATTN,
                    requires_zoom=True,
                    zoom_reasons=["Curved text — visual verification required"],
                    notes=["Curved text — visual verification required" + cross_ref_note],
                    matched_runs=matched_runs,
                    has_3a_flag=has_3a_flag,
                )

            if fuzzy_score >= config.EXACT_MATCH_THRESHOLD:
                status = StatusCode.ATTN if has_3a_flag else (StatusCode.TBD if requires_visual else StatusCode.OK)
                notes = []
                if has_3a_flag:
                    notes.append("Matches copy doc but flagged in A — see copy quality issue")
                else:
                    notes.append("Exact match")
                return MatchFinding(
                    field_name=field_name, panel=panel, language=language,
                    copy_value=copy_text, artwork_value=matched_text,
                    match_type=MatchType.EXACT_MATCH,
                    similarity_score=fuzzy_score, status_code=status,
                    requires_zoom=requires_visual, zoom_reasons=zoom_triggers,
                    notes=notes, matched_runs=matched_runs,
                    has_3a_flag=has_3a_flag,
                )
            elif fuzzy_score >= config.NEAR_MATCH_THRESHOLD:
                notes = [f"Near match ({fuzzy_score:.1f}%) — verify differences{cross_ref_note}"]
                return MatchFinding(
                    field_name=field_name, panel=panel, language=language,
                    copy_value=copy_text, artwork_value=matched_text,
                    match_type=MatchType.NEAR_MATCH,
                    similarity_score=fuzzy_score,
                    status_code=StatusCode.ATTN,
                    requires_zoom=requires_visual, zoom_reasons=zoom_triggers,
                    notes=notes, matched_runs=matched_runs,
                    has_3a_flag=has_3a_flag,
                )
            else:
                notes = [f"Mismatch ({fuzzy_score:.1f}%){cross_ref_note}"]
                return MatchFinding(
                    field_name=field_name, panel=panel, language=language,
                    copy_value=copy_text, artwork_value=matched_text,
                    match_type=MatchType.MISMATCH,
                    similarity_score=fuzzy_score,
                    status_code=StatusCode.FAIL,
                    requires_zoom=requires_visual, zoom_reasons=zoom_triggers,
                    notes=notes, matched_runs=matched_runs,
                    has_3a_flag=has_3a_flag,
                )

        # NOT FOUND
        notes = ["NOT FOUND in artwork — requires visual confirmation"]
        if has_3a_flag:
            notes.append("Note: copy doc also flagged in A")
        return MatchFinding(
            field_name=field_name, panel=panel, language=language,
            copy_value=copy_text, artwork_value=None,
            match_type=MatchType.MISSING_IN_ARTWORK,
            similarity_score=0.0, status_code=StatusCode.FAIL,
            requires_zoom=requires_visual, zoom_reasons=zoom_triggers,
            notes=notes, has_3a_flag=has_3a_flag,
        )

    def _build_text_lookup(self, text_runs: List[TextRun]) -> Dict[str, List[TextRun]]:
        lookup = {}
        for run in text_runs:
            normalized = self.normalizer.normalize(run.text)
            if normalized:
                if normalized not in lookup:
                    lookup[normalized] = []
                lookup[normalized].append(run)
        return lookup

    def _fuzzy_match_sliding_window(
        self, normalized_copy: str, text_runs: List[TextRun]
    ) -> Optional[Tuple[MatchType, float, str, List[TextRun]]]:
        best_score = 0.0
        best_match = None

        # Individual runs
        for run in text_runs:
            normalized_art = self.normalizer.normalize(run.text)
            if not normalized_art:
                continue
            score = SequenceMatcher(None, normalized_copy, normalized_art).ratio() * 100
            if score > best_score:
                best_score = score
                best_match = (normalized_art, [run])

        # Sliding window (2 to 5 runs)
        for window_size in range(2, min(6, len(text_runs) + 1)):
            for i in range(len(text_runs) - window_size + 1):
                window_runs = text_runs[i:i + window_size]
                concatenated = ' '.join(
                    self.normalizer.normalize(r.text)
                    for r in window_runs
                    if self.normalizer.normalize(r.text)
                )
                if not concatenated:
                    continue
                score = SequenceMatcher(None, normalized_copy, concatenated).ratio() * 100
                if score > best_score:
                    best_score = score
                    best_match = (concatenated, window_runs)

        if best_match and best_score >= config.MISMATCH_THRESHOLD:
            matched_text, matched_runs = best_match
            if best_score >= config.EXACT_MATCH_THRESHOLD:
                mt = MatchType.EXACT_MATCH
            elif best_score >= config.NEAR_MATCH_THRESHOLD:
                mt = MatchType.NEAR_MATCH
            else:
                mt = MatchType.MISMATCH
            return (mt, best_score, matched_text, matched_runs)

        return None

    def _log_match_summary(self, findings: List[MatchFinding]):
        exact = sum(1 for f in findings if f.match_type == MatchType.EXACT_MATCH)
        near = sum(1 for f in findings if f.match_type == MatchType.NEAR_MATCH)
        missing = sum(1 for f in findings if f.match_type == MatchType.MISSING_IN_ARTWORK)
        mismatch = sum(1 for f in findings if f.match_type == MatchType.MISMATCH)
        logger.info(f"Match results: {exact} exact, {near} near, {mismatch} mismatch, {missing} missing")


# =============================================================================
# SNAPSHOT GENERATOR (Section 3G)
# =============================================================================

class SnapshotGenerator:
    """Generates annotated full-page snapshots with issue highlights."""

    def generate_snapshots(
        self,
        pdf_path: Path,
        findings: List[MatchFinding],
        output_dir: Path
    ) -> List[Path]:
        if not PIL_AVAILABLE or not PYMUPDF_AVAILABLE:
            logger.warning("PIL/PyMuPDF not available - cannot generate snapshots")
            return []
        if not config.SNAPSHOT_ENABLED:
            return []

        snapshot_paths = []
        try:
            doc = fitz.open(str(pdf_path))
            findings_by_page = self._group_findings_by_page(findings)

            for page_num, page_findings in findings_by_page.items():
                if page_num >= len(doc):
                    continue
                page = doc[page_num]
                pix = page.get_pixmap(dpi=config.SNAPSHOT_DPI)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                img = self._draw_annotations(img, page_findings, page, pix)
                output_dir.mkdir(parents=True, exist_ok=True)
                snapshot_path = output_dir / f"snapshot_page_{page_num + 1}.png"
                img.save(snapshot_path)
                snapshot_paths.append(snapshot_path)
                logger.info(f"Generated snapshot: {snapshot_path}")

            doc.close()
        except Exception as e:
            logger.error(f"Error generating snapshots: {e}")
        return snapshot_paths

    def _group_findings_by_page(self, findings: List[MatchFinding]) -> Dict[int, List[MatchFinding]]:
        by_page = {}
        for finding in findings:
            if not finding.matched_runs:
                continue
            for run in finding.matched_runs:
                if run.bbox:
                    page_num = run.page_number - 1
                    if page_num not in by_page:
                        by_page[page_num] = []
                    by_page[page_num].append(finding)
                    break
        return by_page

    def _draw_annotations(self, img, findings, page, pix):
        draw = ImageDraw.Draw(img)
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 14)
        except:
            font = ImageFont.load_default()

        colors = {
            StatusCode.FAIL: (255, 0, 0),
            StatusCode.ATTN: (255, 165, 0),
            StatusCode.TBD: (255, 255, 0),
            StatusCode.OK: (0, 255, 0),
            StatusCode.FYI: (128, 128, 128),
        }

        for finding in findings:
            if not finding.matched_runs:
                continue
            color = colors.get(finding.status_code, (128, 128, 128))
            for run in finding.matched_runs:
                if not run.bbox:
                    continue
                scale_x = pix.width / page.rect.width
                scale_y = pix.height / page.rect.height
                x0, y0, x1, y1 = run.bbox
                x0i, y0i = int(x0 * scale_x), int(y0 * scale_y)
                x1i, y1i = int(x1 * scale_x), int(y1 * scale_y)
                draw.rectangle([(x0i, y0i), (x1i, y1i)], outline=color, width=3)
                label = finding.issue_id or "?"
                try:
                    bbox = draw.textbbox((x0i, y0i - 20), label, font=font)
                    draw.rectangle(bbox, fill=color)
                    draw.text((x0i, y0i - 20), label, fill=(255, 255, 255), font=font)
                except:
                    draw.text((x0i, y0i - 15), label, fill=color, font=font)
        return img


# =============================================================================
# MARKDOWN RENDERER (Format A Output)
# =============================================================================

class MarkdownRenderer:
    """Renders verification results to markdown in Format A layout."""

    @staticmethod
    def render_complete_report(
        copy_doc: CopyDocument,
        artwork_extraction: ExtractionResult,
        copy_quality_issues: List[CopyQualityIssue],
        claim_risks: List[ClaimRisk],
        conversions: List[ConversionCheck],
        match_findings: List[MatchFinding],
        font_measurements: List[FontMeasurement],
        barcode_results: List[BarcodeResult],
        snapshot_paths: List[str],
    ) -> str:
        sections = []

        sections.append("# Artwork Verification Report\n")
        sections.append(f"*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*\n")
        sections.append(f"*Checker Version: {config.VERSION}*\n")

        # 1. Project Header
        sections.append(MarkdownRenderer._section1(copy_doc))
        # 2. Files
        sections.append(MarkdownRenderer._section2(copy_doc, artwork_extraction))
        # 3. Core Verification Tables
        sections.append("\n## 3️⃣ Core Verification Tables\n")
        sections.append(MarkdownRenderer._section3a(copy_quality_issues))
        sections.append(MarkdownRenderer._section3b(claim_risks))
        sections.append(MarkdownRenderer._section3c(conversions))
        sections.append(MarkdownRenderer._section3d(match_findings, copy_doc))
        sections.append(MarkdownRenderer._section3e(font_measurements))
        sections.append(MarkdownRenderer._section3f(barcode_results))
        sections.append(MarkdownRenderer._section3g(match_findings, snapshot_paths))
        sections.append(MarkdownRenderer._section3h(match_findings, copy_quality_issues, claim_risks, font_measurements))
        # 4. Optional Fields
        sections.append(MarkdownRenderer._section4())
        # 5. Special Notes
        sections.append(MarkdownRenderer._section5(match_findings, artwork_extraction, copy_doc))

        return "\n".join(sections)

    @staticmethod
    def _section1(copy_doc: CopyDocument) -> str:
        return f"""
## 1️⃣ Project Header

| Field | Value |
|-------|-------|
| Project Name | {sanitize_for_markdown(copy_doc.project_name)} |
| Component Type | {sanitize_for_markdown(copy_doc.component_type)} |
"""

    @staticmethod
    def _section2(copy_doc: CopyDocument, artwork: ExtractionResult) -> str:
        art_note = ""
        if artwork.extraction_method == ExtractionMethod.FAILED:
            art_note = "Text extraction failed - visual verification required"
        elif artwork.confidence < 1.0:
            art_note = f"Extraction confidence: {artwork.confidence:.0%}"
        else:
            art_note = f"{len(artwork.text_runs)} text elements extracted"

        copy_note = f"{copy_doc.metadata.get('total_fields', 0)} fields, "
        copy_note += f"{copy_doc.metadata.get('instructional_count', 0)} instructional notes"

        return f"""
## 2️⃣ Files

| Type | Filename | Version | Note |
|------|----------|---------|------|
| Copy Document | {Path(copy_doc.file_path).name} | — | {sanitize_for_markdown(copy_note)} |
| Artwork | {Path(artwork.file_path).name} | — | {sanitize_for_markdown(art_note)} |
"""

    @staticmethod
    def _section3a(issues: List[CopyQualityIssue]) -> str:
        lines = ["\n### A. Copy Quality\n"]
        lines.append("| Language | Field | Issue Type | Recommendation | Status |")
        lines.append("|----------|-------|------------|----------------|--------|")

        if not issues:
            lines.append("| — | All | — | No copy quality issues detected | ✅ |")
        else:
            for issue in issues:
                emoji = config.STATUS_EMOJI.get(issue.status_code.value, "❓")
                lines.append(
                    f"| {issue.language} | "
                    f"{sanitize_for_markdown(truncate_text(issue.field_name, 30))} | "
                    f"{issue.issue_type} | "
                    f"{sanitize_for_markdown(truncate_text(issue.recommendation, 50))} | "
                    f"{emoji} |"
                )
        return "\n".join(lines)

    @staticmethod
    def _section3b(risks: List[ClaimRisk]) -> str:
        lines = ["\n### B. Claim Risk\n"]
        lines.append("| Language | Claim | Risk Level | Rationale | Regions | Action | Status |")
        lines.append("|----------|-------|------------|-----------|---------|--------|--------|")

        if not risks:
            lines.append("| — | No claims detected | — | — | — | — | ✅ |")
        else:
            for risk in risks:
                emoji = config.STATUS_EMOJI.get(risk.status_code.value, "❓")
                lines.append(
                    f"| {risk.language} | "
                    f"{sanitize_for_markdown(truncate_text(risk.claim_text, 40))} | "
                    f"{risk.risk_level.value} | "
                    f"{sanitize_for_markdown(truncate_text(risk.rationale, 40))} | "
                    f"{', '.join(risk.regions[:3])} | "
                    f"{risk.recommended_action} | "
                    f"{emoji} |"
                )
        return "\n".join(lines)

    @staticmethod
    def _section3c(conversions: List[ConversionCheck]) -> str:
        lines = ["\n### C. Label-Claim Conversion\n"]
        lines.append("| Source | Declared (mL) | Calculated (fl oz) | Declared (fl oz) | Within ±0.10 | Status | Notes |")
        lines.append("|--------|---------------|-------------------|------------------|--------------|--------|-------|")

        if not conversions:
            lines.append("| — | — | — | — | — | ✅ | No conversions to check |")
        else:
            for conv in conversions:
                emoji = config.STATUS_EMOJI.get(conv.status_code.value, "❓")
                ml = conv.declared_ml if conv.declared_ml is not None else "-"
                calc = conv.calculated_floz if conv.calculated_floz is not None else "-"
                decl = conv.declared_floz if conv.declared_floz is not None else "-"
                within = "Yes" if conv.within_tolerance else ("No" if conv.within_tolerance is not None else "-")
                lines.append(
                    f"| {conv.source_field} | {ml} | {calc} | {decl} | {within} | {emoji} | {sanitize_for_markdown(conv.notes)} |"
                )
        return "\n".join(lines)

    @staticmethod
    def _section3d(findings: List[MatchFinding], copy_doc: CopyDocument) -> str:
        lines = ["\n### D. Artwork Match\n"]

        # Count deferred instructional fields
        deferred_count = len(copy_doc.instructional_fields)
        if deferred_count > 0:
            # Determine what they're deferred to
            deferred_targets = set()
            for f in copy_doc.instructional_fields:
                if 'details on' in f.text.lower():
                    match = re.search(r'details\s+on\s+(.+)', f.text, re.IGNORECASE)
                    if match:
                        deferred_targets.add(match.group(1).strip())
            if deferred_targets:
                target_str = ", ".join(deferred_targets)
                lines.append(f"*{deferred_count} fields deferred to {target_str} component*\n")
            else:
                lines.append(f"*{deferred_count} fields marked as instructional notes (excluded from matching)*\n")

        # Group by panel and language
        grouped = {}
        for f in findings:
            key = (f.panel, f.language)
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(f)

        for (panel, language), panel_findings in grouped.items():
            lines.append(f"\n**{panel} — {language}**\n")
            lines.append("| Field | Copy Doc Value | Artwork Value | Match | Notes |")
            lines.append("|-------|----------------|---------------|-------|-------|")

            for finding in panel_findings:
                if finding.match_type == MatchType.EXACT_MATCH:
                    match_emoji = "✅" if not finding.has_3a_flag else "⚠️"
                elif finding.match_type == MatchType.NEAR_MATCH:
                    match_emoji = "⚠️"
                elif finding.match_type == MatchType.REQUIRES_VERIFICATION:
                    match_emoji = "🔍"
                else:
                    match_emoji = "❌"

                art_val = finding.artwork_value or "NOT FOUND"
                notes = "; ".join(finding.notes) if finding.notes else ""
                if finding.issue_id:
                    notes = f"[{finding.issue_id}] {notes}"

                lines.append(
                    f"| {sanitize_for_markdown(truncate_text(finding.field_name, 30))} | "
                    f"{sanitize_for_markdown(truncate_text(finding.copy_value, 40))} | "
                    f"{sanitize_for_markdown(truncate_text(art_val, 40))} | "
                    f"{match_emoji} | "
                    f"{sanitize_for_markdown(truncate_text(notes, 60))} |"
                )
        return "\n".join(lines)

    @staticmethod
    def _section3e(measurements: List[FontMeasurement]) -> str:
        lines = ["\n### E. Font Size\n"]
        lines.append("| Text String / Field | Jurisdiction | Required Min (pt) | Measured Min (pt) | Method | Status | Screenshot ID |")
        lines.append("|---------------------|--------------|-------------------|-------------------|--------|--------|---------------|")

        if not measurements:
            lines.append("| — | — | — | — | — | 🔍 Unable to extract font metadata | — |")
        else:
            smallest = min(measurements, key=lambda m: m.font_size_pt)
            usa_pass = smallest.font_size_pt >= 4.5
            eu_pass = smallest.font_size_pt >= 6.0

            lines.append(
                f"| {sanitize_for_markdown(truncate_text(smallest.text, 25))} | "
                f"USA | 4.5 | {smallest.font_size_pt} | Vector | "
                f"{'✅' if usa_pass else '❌'} | FS-001 |"
            )
            lines.append(
                f"| {sanitize_for_markdown(truncate_text(smallest.text, 25))} | "
                f"EU | 6.0 | {smallest.font_size_pt} | Vector | "
                f"{'✅' if eu_pass else '❌'} | FS-001 |"
            )
        return "\n".join(lines)

    @staticmethod
    def _section3f(results: List[BarcodeResult]) -> str:
        lines = ["\n### F. Barcode\n"]
        lines.append("| Symbology | Encoded Digits | Check Digit Valid | X-Dim (mm) | Quiet Zone (mm) | Module Count | Print Contrast | Scan Test |")
        lines.append("|-----------|----------------|-------------------|------------|-----------------|--------------|----------------|-----------|")

        if not results:
            lines.append("| N/A | — | — | — | — | — | — | — |")
        else:
            for result in results:
                check = "✅" if result.check_digit_valid else "❌"
                if not result.scan_successful:
                    check = "—"
                scan = "✅" if result.scan_successful else "Manual check required"
                lines.append(
                    f"| {result.symbology} | "
                    f"{result.decoded_digits or '—'} | "
                    f"{check} | "
                    f"Manual check required | Manual check required | Manual check required | Manual check required | "
                    f"{scan} |"
                )
        return "\n".join(lines)

    @staticmethod
    def _section3g(findings: List[MatchFinding], snapshot_paths: List[str]) -> str:
        lines = ["\n### G. Visual Snapshots\n"]

        # Gather issues
        issue_findings = [f for f in findings if f.issue_id]

        if not issue_findings:
            lines.append("| ID | What | Where | Fix | Linked Rows | Status After Fix |")
            lines.append("|----|------|-------|-----|-------------|------------------|")
            lines.append("| — | No issues flagged | — | — | — | — |")
        else:
            lines.append("| ID | What | Where | Fix | Linked Rows | Status After Fix |")
            lines.append("|----|------|-------|-----|-------------|------------------|")

            for f in issue_findings:
                what = f.match_type.value
                if f.match_type == MatchType.MISSING_IN_ARTWORK:
                    what = "Missing text"
                elif f.match_type == MatchType.MISMATCH:
                    what = "Text mismatch"
                elif f.match_type == MatchType.NEAR_MATCH:
                    what = "Near match"
                elif f.match_type == MatchType.REQUIRES_VERIFICATION:
                    what = "Curved text"

                where = f"{f.panel} — {f.language}"
                fix = "; ".join(f.notes[:1]) if f.notes else "Verify"
                linked = "D"
                if f.has_3a_flag:
                    linked = "A, D"

                lines.append(
                    f"| {f.issue_id} | "
                    f"{sanitize_for_markdown(truncate_text(what, 25))} | "
                    f"{sanitize_for_markdown(truncate_text(where, 25))} | "
                    f"{sanitize_for_markdown(truncate_text(fix, 30))} | "
                    f"{linked} | TBD |"
                )

        if snapshot_paths:
            lines.append(f"\n*{len(snapshot_paths)} annotated snapshot(s) generated.*")
            for path in snapshot_paths:
                lines.append(f"- `{Path(path).name}`")

        return "\n".join(lines)

    @staticmethod
    def _section3h(
        findings: List[MatchFinding],
        quality_issues: List[CopyQualityIssue],
        claim_risks: List[ClaimRisk],
        font_measurements: List[FontMeasurement]
    ) -> str:
        lines = ["\n### H. Score & Summary\n"]
        lines.append("| Area | Checks | Matches | Score % | Notes |")
        lines.append("|------|--------|---------|---------|-------|")

        # Copy accuracy
        total_d = len(findings)
        exact_d = sum(1 for f in findings if f.match_type == MatchType.EXACT_MATCH)
        score_d = (exact_d / total_d * 100) if total_d > 0 else 100
        lines.append(f"| Copy accuracy | {total_d} | {exact_d} | {score_d:.0f} | — |")

        # Claims & risk
        total_b = len(claim_risks)
        ok_b = sum(1 for r in claim_risks if r.status_code == StatusCode.OK)
        score_b = (ok_b / total_b * 100) if total_b > 0 else 100
        claim_note = ""
        high_risk = [r for r in claim_risks if r.risk_level == RiskLevel.HIGH]
        if high_risk:
            claim_note = f"{len(high_risk)} high-risk claim(s) pending"
        lines.append(f"| Claims & risk | {total_b} | {ok_b} | {score_b:.0f} | {claim_note} |")

        # Regulatory (EU font)
        if font_measurements:
            smallest = min(font_measurements, key=lambda m: m.font_size_pt)
            eu_checks = 2  # USA + EU
            eu_pass = 0
            if smallest.font_size_pt >= 4.5:
                eu_pass += 1
            if smallest.font_size_pt >= 6.0:
                eu_pass += 1
            score_e = (eu_pass / eu_checks * 100)
            font_note = f"Smallest: {smallest.font_size_pt}pt"
        else:
            eu_checks = 1
            eu_pass = 0
            score_e = 0
            font_note = "Unable to measure"
        lines.append(f"| Regulatory (EU font) | {eu_checks} | {eu_pass} | {score_e:.0f} | {font_note} |")

        # Overall
        all_scores = [score_d, score_b, score_e]
        overall = sum(all_scores) / len(all_scores)
        lines.append(f"| **Overall** | — | — | **{overall:.1f}** | — |")

        # Top fixes
        lines.append("\n**Top Fixes (❌)**\n")
        lines.append("| Item |")
        lines.append("|------|")
        fail_findings = [f for f in findings if f.status_code == StatusCode.FAIL][:5]
        if fail_findings:
            for f in fail_findings:
                notes_str = "; ".join(f.notes[:1]) if f.notes else f.match_type.value
                lines.append(f"| [{f.issue_id}] {sanitize_for_markdown(f.field_name)}: {sanitize_for_markdown(truncate_text(notes_str, 50))} |")
        else:
            lines.append("| None |")

        # Attention
        lines.append("\n**Attention (⚠️)**\n")
        lines.append("| Item |")
        lines.append("|------|")
        attn_findings = [f for f in findings if f.status_code == StatusCode.ATTN][:5]
        if attn_findings:
            for f in attn_findings:
                notes_str = "; ".join(f.notes[:1]) if f.notes else ""
                lines.append(f"| {sanitize_for_markdown(f.field_name)}: {sanitize_for_markdown(truncate_text(notes_str, 50))} |")
        else:
            lines.append("| None |")

        return "\n".join(lines)

    @staticmethod
    def _section4() -> str:
        return """
## 4️⃣ Optional Fields

| Field | Content |
|-------|---------|
| Version Change Log | Not provided |
| Creative Brand-Voice Check | Not requested |
| One-Page PDF Summary Export | Available on request |
"""

    @staticmethod
    def _section5(findings: List[MatchFinding], artwork: ExtractionResult, copy_doc: CopyDocument) -> str:
        lines = ["\n## 5️⃣ Special Notes / Constraints\n"]
        lines.append("| Constraint | Source | Applies To | Notes |")
        lines.append("|------------|--------|------------|-------|")

        lines.append(
            "| Text must match character-for-character | Brand | All panels | "
            "Including punctuation, case, diacritics |"
        )

        if artwork.extraction_method == ExtractionMethod.FAILED:
            lines.append(
                "| Visual verification required | Technical | All fields | "
                "Text extraction failed (outlined fonts likely) |"
            )

        verify_count = sum(1 for f in findings if f.requires_zoom)
        if verify_count > 0:
            lines.append(
                f"| {verify_count} fields require visual verification | Zoom Triggers | "
                f"See flagged items | Numbers, units, or low confidence |"
            )

        # Note deferred fields
        if copy_doc.instructional_fields:
            deferred = len(copy_doc.instructional_fields)
            lines.append(
                f"| {deferred} instructional notes detected | Copy Document | "
                f"See Section A | Excluded from artwork matching |"
            )

        return "\n".join(lines)


# =============================================================================
# VISION EXPORTER (Pass 1 — --require-vision)
# =============================================================================

class VisionExporter:
    """
    Renders the full artwork page at high DPI and generates focus crops for
    items requiring visual verification.  Evidence is displayed via contact
    sheets (montage images) rather than individual inline images, keeping the
    GPT response stable regardless of item count.

    Two crop sizes are generated per item:
      - focus_crop (tight): standard 100px pad, min 500×250 — used for tiling
      - focus_crop_wide: 300px pad, min 1200×600 — for long/curved/risky text

    Contact sheets (SHEET_COLS × SHEET_ROWS tiles) are created from tight crops
    at 1:1 scale (no downscaling) with an ID label on each tile.
    """

    PAGE_DPI = 300            # Full-page render resolution
    FOCUS_PAD = 100           # Tight crop padding (pixels)
    FOCUS_MIN_W = 500         # Tight crop minimum width
    FOCUS_MIN_H = 250         # Tight crop minimum height
    FOCUS_PAD_WIDE = 300      # Wide crop padding
    FOCUS_MIN_W_WIDE = 1200   # Wide crop minimum width
    FOCUS_MIN_H_WIDE = 600    # Wide crop minimum height
    SHEET_COLS = 3            # Contact sheet columns
    SHEET_ROWS = 3            # Contact sheet rows
    SHEET_MARGIN = 40         # Pixels between tiles and border

    # Field name keywords that always get a wide crop
    _WIDE_CROP_FIELDS = ("ingredient", "marketing copy")

    def _needs_wide_crop(self, finding: "MatchFinding") -> bool:
        """Return True if this finding warrants a wide context crop."""
        if any("curved" in r.lower() for r in finding.zoom_reasons):
            return True
        if finding.match_type in (
            MatchType.NEAR_MATCH,
            MatchType.MISMATCH,
            MatchType.MISSING_IN_ARTWORK,
        ):
            return True
        if len(finding.copy_value or "") > 80:
            return True
        fn_lower = (finding.field_name or "").lower()
        if any(kw in fn_lower for kw in self._WIDE_CROP_FIELDS):
            return True
        return False

    def _make_crop(
        self,
        full_img: "Image.Image",
        run_bboxes: List[Tuple],
        scale: float,
        pad: int,
        min_w: int,
        min_h: int,
    ) -> "Image.Image":
        """Crop full_img to the union bbox of run_bboxes, expanded by pad/min dims."""
        img_w, img_h = full_img.size
        ux0 = min(b[0] for b in run_bboxes) * scale
        uy0 = min(b[1] for b in run_bboxes) * scale
        ux1 = max(b[2] for b in run_bboxes) * scale
        uy1 = max(b[3] for b in run_bboxes) * scale
        cx = (ux0 + ux1) / 2
        cy = (uy0 + uy1) / 2
        half_w = max((ux1 - ux0) / 2 + pad, min_w / 2)
        half_h = max((uy1 - uy0) / 2 + pad, min_h / 2)
        x0 = max(0, int(cx - half_w))
        y0 = max(0, int(cy - half_h))
        x1 = min(img_w, int(cx + half_w))
        y1 = min(img_h, int(cy + half_h))
        return full_img.crop((x0, y0, x1, y1))

    def _generate_contact_sheets(
        self,
        items: List[Dict[str, Any]],
        sheets_dir: Path,
    ) -> List[Dict[str, Any]]:
        """
        Create contact sheet montages from tight focus crops at 1:1 scale.
        Returns list of {"sheet_path": ..., "items": [...]} dicts.
        """
        sheets_dir.mkdir(parents=True, exist_ok=True)
        croppable = [it for it in items if it.get("focus_crop")]
        if not croppable:
            return []

        per_sheet = self.SHEET_COLS * self.SHEET_ROWS
        contact_sheets = []
        sheet_idx = 1
        label_h = 32  # pixels for ID label bar at top of each tile

        for batch_start in range(0, len(croppable), per_sheet):
            batch = croppable[batch_start: batch_start + per_sheet]
            loaded: List[Tuple[str, "Image.Image"]] = []
            for it in batch:
                try:
                    img = Image.open(it["focus_crop"]).convert("RGB")
                    loaded.append((it["id"], img))
                except Exception:
                    continue
            if not loaded:
                sheet_idx += 1
                continue

            cell_w = max(img.width for _, img in loaded)
            cell_h = max(img.height for _, img in loaded)
            tile_h = cell_h + label_h
            m = self.SHEET_MARGIN
            cols = min(self.SHEET_COLS, len(loaded))
            rows = math.ceil(len(loaded) / cols)

            canvas_w = cols * cell_w + (cols + 1) * m
            canvas_h = rows * tile_h + (rows + 1) * m
            canvas = Image.new("RGB", (canvas_w, canvas_h), (255, 255, 255))
            draw = ImageDraw.Draw(canvas)
            try:
                font = ImageFont.truetype("arial.ttf", 20)
            except Exception:
                try:
                    font = ImageFont.truetype(
                        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 20
                    )
                except Exception:
                    font = ImageFont.load_default()

            ids_in_sheet: List[str] = []
            for idx, (item_id, img) in enumerate(loaded):
                col = idx % cols
                row = idx // cols
                tile_x = m + col * (cell_w + m)
                tile_y = m + row * (tile_h + m)
                # Dark label bar
                draw.rectangle(
                    [tile_x, tile_y, tile_x + cell_w, tile_y + label_h],
                    fill=(40, 40, 40),
                )
                draw.text((tile_x + 6, tile_y + 6), item_id, fill=(255, 255, 255), font=font)
                # Paste crop at 1:1 below label
                canvas.paste(img, (tile_x, tile_y + label_h))
                ids_in_sheet.append(item_id)

            sheet_path = sheets_dir / f"sheet_{sheet_idx:02d}.png"
            canvas.save(str(sheet_path))
            logger.info(f"Contact sheet {sheet_idx}: {len(loaded)} crops → {sheet_path}")
            contact_sheets.append({"sheet_path": str(sheet_path), "items": ids_in_sheet})
            sheet_idx += 1

        return contact_sheets

    def export(
        self,
        pdf_path: Path,
        match_findings: List[MatchFinding],
        output_dir: Path,
    ) -> Path:
        """Render full-page images, focus crops, wide crops, contact sheets;
        returns vision_tasks.json path."""
        if not PIL_AVAILABLE or not PYMUPDF_AVAILABLE:
            raise RuntimeError("PIL and PyMuPDF must be installed for vision export")

        vision_dir = output_dir / "gpt_vision"
        pages_dir = vision_dir / "pages"
        crops_dir = vision_dir / "crops"
        wide_dir = vision_dir / "focus_wide"
        sheets_dir = vision_dir / "sheets"
        for d in (vision_dir, pages_dir, crops_dir, wide_dir):
            d.mkdir(parents=True, exist_ok=True)

        doc = fitz.open(str(pdf_path))
        page_images: Dict[int, Image.Image] = {}
        page_image_paths: Dict[int, str] = {}

        for page_idx in range(len(doc)):
            page = doc[page_idx]
            page_num = page_idx + 1
            mat = fitz.Matrix(self.PAGE_DPI / 72, self.PAGE_DPI / 72)
            pix = page.get_pixmap(matrix=mat)
            full_img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            page_img_path = pages_dir / f"page_{page_num}.png"
            full_img.save(str(page_img_path))
            page_images[page_num] = full_img
            page_image_paths[page_num] = str(page_img_path)
            logger.info(f"Rendered full page {page_num} at {self.PAGE_DPI} DPI: {page_img_path}")

        doc.close()

        # --- Build vision items with tight + optional wide crops ---
        scale = self.PAGE_DPI / 72
        items = []
        for finding in match_findings:
            needs_vision = (
                finding.requires_zoom
                or finding.status_code in (StatusCode.FAIL, StatusCode.ATTN, StatusCode.TBD)
            )
            if not needs_vision:
                continue

            page_num = 1
            if finding.matched_runs:
                for run in finding.matched_runs:
                    if run.page_number:
                        page_num = run.page_number
                        break

            _fallback_id = f"{finding.field_name}|{finding.panel}|{finding.language}"
            safe_id = (finding.issue_id or _fallback_id).replace("/", "_").replace("|", "_")

            run_bboxes = [
                r.bbox for r in finding.matched_runs
                if r.bbox and r.page_number == page_num
            ]

            focus_crop_path: Optional[str] = None
            focus_crop_wide_path: Optional[str] = None

            if run_bboxes and page_num in page_images:
                full_img = page_images[page_num]

                # Tight crop
                crop = self._make_crop(
                    full_img, run_bboxes, scale,
                    self.FOCUS_PAD, self.FOCUS_MIN_W, self.FOCUS_MIN_H,
                )
                crop_path = crops_dir / f"{safe_id}_page_{page_num}.png"
                crop.save(str(crop_path))
                focus_crop_path = str(crop_path)

                # Wide crop for risky items
                if self._needs_wide_crop(finding):
                    wide = self._make_crop(
                        full_img, run_bboxes, scale,
                        self.FOCUS_PAD_WIDE, self.FOCUS_MIN_W_WIDE, self.FOCUS_MIN_H_WIDE,
                    )
                    wide_path = wide_dir / f"{safe_id}_page_{page_num}.png"
                    wide.save(str(wide_path))
                    focus_crop_wide_path = str(wide_path)

            item: Dict[str, Any] = {
                "id": finding.issue_id or _fallback_id,
                "panel": finding.panel,
                "language": finding.language,
                "field_name": finding.field_name,
                "copy_value": finding.copy_value,
                "script_artwork_value": finding.artwork_value or "NOT FOUND",
                "script_match_type": finding.match_type.value,
                "status_code": finding.status_code.value,
                "zoom_reasons": finding.zoom_reasons,
                "page_guess": page_num,
                "page_image": page_image_paths.get(page_num, ""),
                "focus_crop": focus_crop_path,
                "focus_crop_wide": focus_crop_wide_path,
            }
            items.append(item)

        # --- Generate contact sheets ---
        contact_sheets = self._generate_contact_sheets(items, sheets_dir)

        # --- Write vision_tasks.json ---
        required_ids = [it["id"] for it in items]
        hash_payload_items = []
        for it in items:
            hash_payload_items.append({
                "id": it.get("id"),
                "copy_value": it.get("copy_value", ""),
                "script_artwork_value": it.get("script_artwork_value", ""),
                "script_match_type": it.get("script_match_type", ""),
                "status_code": it.get("status_code", ""),
                "zoom_reasons": it.get("zoom_reasons", []),
                "page_guess": it.get("page_guess", 1),
                "focus_crop": bool(it.get("focus_crop")),
            })
        hash_payload = json.dumps(hash_payload_items, ensure_ascii=False, sort_keys=True).encode("utf-8")
        task_hash = hashlib.sha256(hash_payload).hexdigest()

        tasks_data: Dict[str, Any] = {
            "enabled": True,
            "artwork_pdf": str(pdf_path),
            "output_dir": str(output_dir),
            "page_images": {str(k): v for k, v in page_image_paths.items()},
            "contact_sheets": contact_sheets,
            "required_ids": required_ids,
            "task_hash": task_hash,
            "items": items,
        }
        tasks_path = vision_dir / "vision_tasks.json"
        tasks_path.write_text(json.dumps(tasks_data, indent=2, ensure_ascii=False), encoding="utf-8")
        logger.info(f"Vision tasks written: {tasks_path} ({len(items)} items, "
                    f"{len(contact_sheets)} sheet(s))")

        # --- Write VISION_INSTRUCTIONS.md ---
        instructions = (
            "# GPT Vision Verification Required\n\n"
            "This run produced items requiring visual verification.\n\n"
            "## Workflow\n\n"
            "1. Open `vision_tasks.json`\n"
            "2. View the full page image (orientation) then contact sheet(s) (zoomed tiles).\n"
            "3. For each item in the item list:\n"
            "   - Read the tile labeled with the item ID on the contact sheet.\n"
            "   - For wide-crop items, open `focus_crop_wide` for long/curved text.\n"
            "   - `NOT FOUND` items have no crop — scan the full page image.\n"
            "   - Retype the ACTUAL visible artwork text character-by-character.\n"
            "4. Copy the HUMAN TOKEN printed to stdout by Pass 1.\n"
            "   It is also stored in `output/.HUMAN_TOKEN`.\n"
            "5. Write `output/vision_overrides.json` (start from vision_overrides.prefill.json):\n\n"
            "```json\n"
            '{"vision_audit":{\n'
            '    "source":"manual_image_read",\n'
            '    "task_hash":"<task_hash from vision_tasks.json>",\n'
            '    "human_token":"<token from stdout / .HUMAN_TOKEN>"\n'
            '  },\n'
            '"overrides":[\n'
            '  {"finding_id":"<id>","visual_artwork_value":"<text>","found":true,\n'
            '   "confirmed":true,"notes":"Visually verified on page X, [panel], [lang]",\n'
            '   "evidence":"focus_crop_wide|focus_crop|page_image","evidence_path":"<path>"}\n'
            "]}\n"
            "```\n\n"
            "Sentinel block is printed to stdout between:\n"
            "`<<<GPT_VISION_REQUIRED>>>` and `<<<END_GPT_VISION_REQUIRED>>>`\n"
        )
        (vision_dir / "VISION_INSTRUCTIONS.md").write_text(instructions, encoding="utf-8")

        return tasks_path

    @staticmethod
    def print_sentinel(tasks_path: Path, item_count: int) -> None:
        """Print the GPT_VISION_REQUIRED sentinel block to stdout."""
        print("<<<GPT_VISION_REQUIRED>>>")
        print(json.dumps({
            "vision_tasks": str(tasks_path),
            "item_count": item_count,
            "instructions": (
                "1) Open vision_tasks.json. "
                "2) View page image (orientation) then contact sheet(s) (zoomed tiles). "
                "3) For each item: read its labeled tile; use focus_crop_wide for long/curved text. "
                "NOT FOUND items — scan full page_image. "
                "4) Retype actual visible artwork text character-by-character. "
                "5) Write output/vision_overrides.json. "
                "6) Re-run with --vision-overrides ./output/vision_overrides.json"
            ),
        }, indent=2))
        print("<<<END_GPT_VISION_REQUIRED>>>")


# =============================================================================
# VISION OVERRIDE APPLIER (Pass 2 — --vision-overrides)
# =============================================================================

class VisionOverrideApplier:
    """
    Reads vision_overrides.json produced by the GPT visual pass and patches
    match_findings in-place before the final report is rendered.
    """

    normalizer = TextNormalizer

    def _load_tasks_manifest(self, overrides_path: Path) -> Dict[str, Any]:
        """Load vision_tasks.json produced in Pass 1 (sibling gpt_vision/ dir)."""
        tasks_path = overrides_path.parent / "gpt_vision" / "vision_tasks.json"
        if not tasks_path.exists():
            raise FileNotFoundError(
                f"vision_tasks.json not found for validation: {tasks_path}. "
                "Run Pass 1 first to generate tasks."
            )
        return json.loads(tasks_path.read_text(encoding="utf-8"))

    def _validate_overrides(self, tasks: Dict[str, Any], raw_overrides: Dict[str, Any]) -> None:
        if not config.VISION_OVERRIDE_STRICT:
            return

        required_ids = set(tasks.get("required_ids", []))
        task_hash = tasks.get("task_hash", "")
        overrides = raw_overrides.get("overrides", [])

        # 1) Audit stamp
        if config.VISION_OVERRIDE_REQUIRE_AUDIT_STAMP:
            audit = raw_overrides.get("vision_audit")
            if not isinstance(audit, dict) or audit.get("source") != "manual_image_read":
                raise VisionOverrideRejected(
                    "VISION OVERRIDE REJECTED: Missing/invalid vision_audit stamp."
                )
            if task_hash and audit.get("task_hash") != task_hash:
                raise VisionOverrideRejected(
                    "VISION OVERRIDE REJECTED: task_hash mismatch. "
                    "Overrides do not match current vision_tasks.json."
                )

        # 2) IDs: no duplicates, no missing, no extras
        seen = [ov.get("finding_id", "") for ov in overrides]
        if len(seen) != len(set(seen)):
            raise VisionOverrideRejected(
                "VISION OVERRIDE REJECTED: Duplicate finding_id entries in overrides."
            )
        seen_set = set(seen)
        missing = required_ids - seen_set
        extra = seen_set - required_ids
        if missing:
            raise VisionOverrideRejected(
                f"VISION OVERRIDE REJECTED: Missing overrides for required IDs: {sorted(missing)[:10]}"
            )
        if extra:
            raise VisionOverrideRejected(
                f"VISION OVERRIDE REJECTED: Overrides include unknown IDs: {sorted(extra)[:10]}"
            )

        # 3) Evidence required per item
        if config.VISION_OVERRIDE_REQUIRE_EVIDENCE:
            for ov in overrides:
                fid = ov.get("finding_id", "")
                found = bool(ov.get("found", False))
                v = (ov.get("visual_artwork_value") or "").strip()
                ev = (ov.get("evidence") or "").strip()
                evp = (ov.get("evidence_path") or "").strip()
                if found and not v:
                    raise VisionOverrideRejected(
                        f"VISION OVERRIDE REJECTED: '{fid}' marked found=true but visual_artwork_value is empty."
                    )
                if not (ev or evp):
                    raise VisionOverrideRejected(
                        f"VISION OVERRIDE REJECTED: '{fid}' missing evidence/evidence_path."
                    )
                # Validate evidence_path actually exists on disk (within output dir)
                if evp:
                    evp_path = Path(evp)
                    if not evp_path.is_absolute():
                        evp_path = Path(tasks.get("output_dir", ".")) / evp_path
                    if not evp_path.exists():
                        raise VisionOverrideRejected(
                            f"VISION OVERRIDE REJECTED: '{fid}' evidence_path does not exist: {evp}"
                        )

        # Note: identical-rate check removed — when the script extracts text
        # correctly and the GPT confirms it visually, values will legitimately
        # match. The audit stamp + evidence fields are the authoritative gate.

    def apply(
        self,
        findings: List[MatchFinding],
        overrides_path: Path,
    ) -> None:
        """Mutates findings in-place based on vision overrides."""
        if not overrides_path.exists():
            raise FileNotFoundError(f"vision_overrides.json not found: {overrides_path}")

        raw = json.loads(overrides_path.read_text(encoding="utf-8"))
        tasks = self._load_tasks_manifest(overrides_path)
        self._validate_overrides(tasks, raw)
        overrides: List[Dict[str, Any]] = raw.get("overrides", [])

        # Build id → finding index map (by issue_id, with composite fallback for no-issue_id items)
        id_map: Dict[str, int] = {}
        for idx, f in enumerate(findings):
            if f.issue_id:
                id_map[f.issue_id] = idx
            else:
                fallback = f"{f.field_name}|{f.panel}|{f.language}"
                id_map[fallback] = idx

        applied = 0
        for ov in overrides:
            fid = ov.get("finding_id", "")
            if fid not in id_map:
                logger.warning(f"Vision override finding_id '{fid}' not found in findings")
                continue

            finding = findings[id_map[fid]]
            found: bool = ov.get("found", False)
            visual_value: str = ov.get("visual_artwork_value", "")
            notes_override: str = ov.get("notes", f"Visually verified")

            if not found or not visual_value:
                # Still missing after visual check
                finding.notes = [notes_override] if notes_override else finding.notes
                finding.status_code = StatusCode.FAIL
                finding.requires_zoom = False
                applied += 1
                continue

            # Update artwork value
            finding.artwork_value = visual_value
            finding.requires_zoom = False

            # Recalculate match quality
            norm_copy = self.normalizer.normalize(finding.copy_value)
            norm_art = self.normalizer.normalize(visual_value)

            if norm_copy == norm_art:
                finding.match_type = MatchType.EXACT_MATCH
                finding.similarity_score = 100.0
                finding.status_code = StatusCode.OK if not finding.has_3a_flag else StatusCode.ATTN
            else:
                score = SequenceMatcher(None, norm_copy, norm_art).ratio() * 100
                finding.similarity_score = score
                if score >= Config.EXACT_MATCH_THRESHOLD:
                    finding.match_type = MatchType.EXACT_MATCH
                    finding.status_code = StatusCode.OK if not finding.has_3a_flag else StatusCode.ATTN
                elif score >= Config.NEAR_MATCH_THRESHOLD:
                    finding.match_type = MatchType.NEAR_MATCH
                    finding.status_code = StatusCode.ATTN
                else:
                    finding.match_type = MatchType.MISMATCH
                    finding.status_code = StatusCode.FAIL

            finding.notes = [notes_override]
            applied += 1

        logger.info(f"Applied {applied}/{len(overrides)} vision overrides")


# =============================================================================
# SHARED OVERRIDE VALIDATION (used by Pass 2 and run_artwork_check.py)
# =============================================================================

def _load_vision_tasks_manifest(output_dir: str) -> Dict[str, Any]:
    """Load Pass 1 vision task manifest (vision_tasks.json) if it exists."""
    tasks_path = Path(output_dir) / "gpt_vision" / "vision_tasks.json"
    if not tasks_path.exists():
        raise FileNotFoundError(
            f"vision_tasks.json not found: {tasks_path}. Run Pass 1 first."
        )
    return json.loads(tasks_path.read_text(encoding="utf-8"))


def validate_vision_overrides(
    output_dir: str,
    vision_overrides_path: str,
    check_human_token: bool = True,
) -> None:
    """
    Structural gate: refuse Pass 2 unless vision_overrides.json contains a valid
    vision_audit stamp, matching task_hash, human_token (when check_human_token=True),
    per-item evidence, confirmed flags, and no missing/extra IDs.

    Raises ValueError on any validation failure.
    """
    tasks = _load_vision_tasks_manifest(output_dir)
    ov_path = Path(vision_overrides_path)
    raw = json.loads(ov_path.read_text(encoding="utf-8"))

    # 1. Audit stamp
    audit = raw.get("vision_audit")
    if not isinstance(audit, dict) or audit.get("source") != "manual_image_read":
        raise ValueError(
            "VISION OVERRIDE REJECTED: Missing/invalid vision_audit stamp "
            "(source must be 'manual_image_read')."
        )

    # 2. task_hash
    task_hash = tasks.get("task_hash")
    if task_hash and audit.get("task_hash") != task_hash:
        raise ValueError(
            "VISION OVERRIDE REJECTED: task_hash mismatch vs current vision_tasks.json."
        )

    # 3. human_token (Pass 2 structural gate)
    if check_human_token:
        token_path = Path(output_dir) / ".HUMAN_TOKEN"
        if token_path.exists():
            expected_token = token_path.read_text(encoding="utf-8").strip()
            provided_token = (audit.get("human_token") or "").strip()
            if provided_token != expected_token:
                raise ValueError(
                    "VISION OVERRIDE REJECTED: human_token mismatch. "
                    f"Expected token from output/.HUMAN_TOKEN does not match "
                    f"vision_audit.human_token in overrides file."
                )
        else:
            raise ValueError(
                "VISION OVERRIDE REJECTED: output/.HUMAN_TOKEN not found. "
                "Run Pass 1 to generate the token."
            )

    # 4. overrides list + ID coverage
    overrides = raw.get("overrides", [])
    if not isinstance(overrides, list):
        raise ValueError("VISION OVERRIDE REJECTED: overrides must be a list.")
    required_ids = set(tasks.get("required_ids", []))
    if not overrides and required_ids:
        raise ValueError(
            "VISION OVERRIDE REJECTED: overrides list is empty but "
            f"{len(required_ids)} vision item(s) require confirmation."
        )

    # 5. ID coverage — no duplicates, no missing, no extras
    seen = [o.get("finding_id") for o in overrides]
    if len(seen) != len(set(seen)):
        raise ValueError("VISION OVERRIDE REJECTED: duplicate finding_id in overrides.")
    missing = required_ids - set(seen)
    extra = set(seen) - required_ids
    if missing:
        raise ValueError(
            f"VISION OVERRIDE REJECTED: missing required IDs: {sorted(missing)[:10]}"
        )
    if extra:
        raise ValueError(
            f"VISION OVERRIDE REJECTED: unknown IDs present: {sorted(extra)[:10]}"
        )

    # 6. Per-item: confirmed, found→value, evidence
    for o in overrides:
        fid = o.get("finding_id", "?")
        found = bool(o.get("found", False))
        v = (o.get("visual_artwork_value") or "").strip()
        ev = (o.get("evidence") or "").strip()
        evp = (o.get("evidence_path") or "").strip()

        if o.get("confirmed") is not True:
            raise ValueError(
                f"VISION OVERRIDE REJECTED: {fid} must have confirmed: true."
            )
        if found and not v:
            raise ValueError(
                f"VISION OVERRIDE REJECTED: {fid} found=true but visual_artwork_value empty."
            )
        if not (ev or evp):
            raise ValueError(
                f"VISION OVERRIDE REJECTED: {fid} missing evidence/evidence_path."
            )
        if evp:
            evp_path = Path(evp)
            if not evp_path.is_absolute():
                evp_path = Path(output_dir) / evp_path
            if not evp_path.exists():
                raise ValueError(
                    f"VISION OVERRIDE REJECTED: {fid} evidence_path does not exist: {evp}"
                )


